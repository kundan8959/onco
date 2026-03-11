#!/usr/bin/env python3
"""
Generate Oncology EHR Project Workplan Excel Workbook
Upload this .xlsx to Google Sheets → File → Import → Upload
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

# ── Shared Styles ─────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
SUB_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="1F2937")
BODY_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="FFFFFF")

DARK_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
GREEN_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
TEAL_FILL = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")

WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def style_header_row(ws, row, num_cols, fill=HEADER_FILL):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_body(ws, start_row, end_row, num_cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if r % 2 == 0:
                cell.fill = LIGHT_FILL


def auto_width(ws, num_cols, min_w=15, max_w=50):
    for col in range(1, num_cols + 1):
        max_len = min_w
        letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = max_len + 2


# ═══════════════════════════════════════════════════════════════════
# TAB 1 — PROJECT OVERVIEW
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Project Overview"
ws1.sheet_properties.tabColor = "1E3A5F"

overview_data = [
    ["Section", "Details"],
    ["Project Name", "Oncology Care Management EHR Platform"],
    ["Project Type", "SaaS — Hospital & Patient-Facing EHR System for Oncology"],
    ["Prepared For", "[Client / Hospital Name]"],
    ["Prepared By", "TPAI Engineering Team"],
    ["Date", "March 2026"],
    ["Version", "1.0"],
    ["", ""],
    ["VISION", ""],
    ["Purpose",
     "Build a comprehensive, role-based oncology EHR platform that hospitals can deploy to manage cancer patients through their entire treatment journey — from diagnosis through chemotherapy cycles, emergency events, medication management, and ongoing benchmarking."],
    ["Patient Value",
     "Patients get a personal portal to track remaining chemo sessions, view treatment progress, receive notifications for missed appointments, and stay connected with their care team."],
    ["Hospital Value",
     "Hospitals get a powerful dashboard with real-time benchmarks, compliance tracking, ER visit monitoring, role-based workflows for doctors/nurses/chemists, and automated notification systems."],
    ["", ""],
    ["CANCER TYPES COVERED", ""],
    ["1", "Breast Cancer — Chemotherapy protocols, hormone therapy tracking, HER2 status monitoring"],
    ["2", "Prostate Cancer — PSA benchmarking, radiation + chemo cycles, hormone therapy tracking"],
    ["3", "Lung Cancer — Staging (SCLC/NSCLC), immunotherapy + chemo tracking, PFT benchmarks"],
    ["4", "Colorectal Cancer — CEA marker tracking, surgical + chemo timelines, colonoscopy scheduling"],
    ["", ""],
    ["KEY CAPABILITIES", ""],
    ["Chemo Tracking", "Track total cycles prescribed vs completed, next session dates, missed sessions with reasons"],
    ["Smart Notifications", "Automated alerts via Email/SMS/Push to patient, emergency contact, and care team"],
    ["Benchmarking", "Tumor marker trends (up/down), treatment response rates, survival benchmarks per cancer type"],
    ["ER Monitoring", "Track emergency room visits, extra medications administered, unplanned hospitalizations"],
    ["Role-Based Access", "SuperAdmin, Hospital Admin, Doctor, Nurse, Chemist/Pharmacist, Patient — each with specific permissions"],
    ["Reporting", "Exportable reports for compliance, insurance, and clinical audits"],
]

for r, row in enumerate(overview_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = DARK_FILL
        elif val in ("VISION", "CANCER TYPES COVERED", "KEY CAPABILITIES"):
            cell.font = Font(name="Calibri", bold=True, size=12, color="1E3A5F")
            cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            ws1.cell(row=r, column=2).fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        else:
            cell.font = BODY_FONT

ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 90


# ═══════════════════════════════════════════════════════════════════
# TAB 2 — PHASE-WISE PROJECT PLAN
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Phase-Wise Plan")
ws2.sheet_properties.tabColor = "2563EB"

phases = [
    ["Phase", "Sprint", "Duration", "Module / Deliverable", "Key Tasks", "Owner", "Status"],
    # Phase 1
    ["Phase 1: Foundation & Auth", "Sprint 1", "Week 1–2",
     "Project Setup & Infrastructure",
     "• Monorepo setup (React + NestJS)\n• Database schema design (PostgreSQL)\n• CI/CD pipeline (GitHub Actions)\n• Docker containerization\n• Environment configs (Dev/Staging/Prod)",
     "Full Stack Team", "Not Started"],
    ["Phase 1: Foundation & Auth", "Sprint 2", "Week 3–4",
     "Authentication & Role Management",
     "• JWT-based authentication\n• Role-based access control (RBAC)\n• Roles: SuperAdmin, Hospital Admin, Doctor, Nurse, Chemist, Patient\n• Login/Register/Forgot Password flows\n• Session management & token refresh\n• Audit logging for all auth events",
     "Backend Lead", "Not Started"],
    ["Phase 1: Foundation & Auth", "Sprint 2", "Week 3–4",
     "User Management Module",
     "• SuperAdmin: CRUD for hospitals, admins\n• Hospital Admin: CRUD for doctors, nurses, chemists\n• Profile management for all roles\n• Role permission matrix implementation",
     "Full Stack Team", "Not Started"],

    # Phase 2
    ["Phase 2: Core Patient & Oncology", "Sprint 3", "Week 5–6",
     "Patient Registration & Profile",
     "• Patient demographics & medical history\n• Insurance & emergency contact management\n• Cancer diagnosis intake form\n• Document upload (reports, scans)\n• Patient search & filtering",
     "Backend + Frontend", "Not Started"],
    ["Phase 2: Core Patient & Oncology", "Sprint 4", "Week 7–8",
     "Breast Cancer Module",
     "• Diagnosis details (ER/PR/HER2 status, tumor grade)\n• Chemotherapy protocol assignment (AC-T, TC, etc.)\n• Cycle tracking (prescribed vs completed)\n• Hormone therapy tracking\n• Tumor marker benchmarks (CA 15-3)\n• Treatment response assessment",
     "Oncology Team", "Not Started"],
    ["Phase 2: Core Patient & Oncology", "Sprint 5", "Week 9–10",
     "Prostate Cancer Module",
     "• Diagnosis (Gleason score, PSA levels, staging)\n• Chemo protocol assignment (Docetaxel, Cabazitaxel)\n• PSA trend benchmarking (rising/falling)\n• Hormone/ADT therapy tracking\n• Radiation therapy integration\n• Treatment response metrics",
     "Oncology Team", "Not Started"],
    ["Phase 2: Core Patient & Oncology", "Sprint 6", "Week 11–12",
     "Lung Cancer Module",
     "• Diagnosis (SCLC vs NSCLC, staging, mutations)\n• Chemo protocol (Cisplatin/Carboplatin combos)\n• Immunotherapy tracking (Pembrolizumab, etc.)\n• PFT (Pulmonary Function Test) benchmarks\n• Tumor marker monitoring\n• Treatment line tracking (1st, 2nd, 3rd line)",
     "Oncology Team", "Not Started"],
    ["Phase 2: Core Patient & Oncology", "Sprint 7", "Week 13–14",
     "Colorectal Cancer Module",
     "• Diagnosis (staging, tumor location, MSI status)\n• Chemo protocol (FOLFOX, FOLFIRI, CAPOX)\n• CEA marker trend tracking\n• Surgical timeline integration\n• Colonoscopy scheduling & follow-up\n• Treatment response (RECIST criteria)",
     "Oncology Team", "Not Started"],

    # Phase 3
    ["Phase 3: Chemo & Treatment Engine", "Sprint 8", "Week 15–16",
     "Chemotherapy Scheduling Engine",
     "• Calendar-based chemo cycle scheduling\n• Auto-calculate next session dates\n• Missed session detection & flagging\n• Pre-chemo checklist (blood work, vitals)\n• Cycle completion tracking (X of Y done)\n• Treatment delay documentation",
     "Backend Lead", "Not Started"],
    ["Phase 3: Chemo & Treatment Engine", "Sprint 8", "Week 15–16",
     "Medication & Prescription Management",
     "• Chemo drug catalog (per cancer type)\n• Dosage calculation (BSA-based)\n• Supportive medication tracking (anti-nausea, etc.)\n• Extra/emergency medication logging\n• Drug interaction alerts\n• Chemist/Pharmacist verification workflow",
     "Backend + Chemist Flow", "Not Started"],

    # Phase 4
    ["Phase 4: Notifications & Alerts", "Sprint 9", "Week 17–18",
     "Notification System",
     "• Email notifications (SendGrid/AWS SES)\n• SMS notifications (Twilio)\n• Push notifications (Firebase)\n• Notification templates per event type\n• Notification preferences per user\n• Notification history & read tracking",
     "Backend Team", "Not Started"],
    ["Phase 4: Notifications & Alerts", "Sprint 9", "Week 17–18",
     "Alert Rules Engine",
     "• Missed chemo session → Notify patient + emergency contact\n• Doctor unavailable → Notify patient + assign backup\n• Abnormal benchmark → Alert care team\n• ER visit logged → Notify oncologist\n• Treatment milestone → Notify patient\n• Custom alert rules per hospital",
     "Backend Team", "Not Started"],

    # Phase 5
    ["Phase 5: Benchmarks & Analytics", "Sprint 10", "Week 19–20",
     "Benchmarking Dashboard",
     "• Tumor marker trend charts (up/down arrows)\n• Treatment response rates per cancer type\n• Chemo completion rates\n• Side-effect frequency tracking\n• Comparative benchmarks (hospital vs national)\n• Patient outcome scoring",
     "Frontend + Data", "Not Started"],
    ["Phase 5: Benchmarks & Analytics", "Sprint 10", "Week 19–20",
     "ER & Emergency Tracking",
     "• Emergency room visit logging\n• Reason categorization (fever, pain, reaction, etc.)\n• Extra medications administered in ER\n• Unplanned hospitalization tracking\n• ER visit trends per patient/cancer type\n• Cost impact analysis",
     "Full Stack Team", "Not Started"],

    # Phase 6
    ["Phase 6: Patient Portal", "Sprint 11", "Week 21–22",
     "Patient-Facing Portal",
     "• Treatment journey timeline view\n• Remaining chemo sessions counter\n• Next appointment details\n• Medication list & schedule\n• Doctor/nurse contact info\n• Symptom self-reporting\n• Document access (reports, prescriptions)",
     "Frontend Team", "Not Started"],
    ["Phase 6: Patient Portal", "Sprint 11", "Week 21–22",
     "Patient Notifications & Communication",
     "• Appointment reminders (email + SMS + push)\n• Missed session follow-up alerts\n• Treatment summary after each cycle\n• Educational content per cancer type\n• Secure messaging with care team",
     "Full Stack Team", "Not Started"],

    # Phase 7
    ["Phase 7: Reporting & Compliance", "Sprint 12", "Week 23–24",
     "Reports & Exports",
     "• Patient treatment summary reports (PDF)\n• Hospital-wide oncology analytics\n• Insurance/billing report generation\n• Compliance audit reports\n• Custom report builder\n• Scheduled report delivery (email)",
     "Full Stack Team", "Not Started"],

    # Phase 8
    ["Phase 8: Testing & Deployment", "Sprint 13–14", "Week 25–28",
     "QA, UAT & Deployment",
     "• Unit & integration testing (>80% coverage)\n• End-to-end testing (Cypress/Playwright)\n• Security audit & penetration testing\n• HIPAA compliance review\n• Performance & load testing\n• UAT with hospital staff\n• Production deployment\n• Documentation & training materials",
     "QA + DevOps", "Not Started"],
]

for r, row in enumerate(phases, 1):
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = DARK_FILL
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 35
ws2.column_dimensions["E"].width = 60
ws2.column_dimensions["F"].width = 22
ws2.column_dimensions["G"].width = 14
ws2.row_dimensions[1].height = 30
for r in range(2, len(phases) + 1):
    ws2.row_dimensions[r].height = 100


# ═══════════════════════════════════════════════════════════════════
# TAB 3 — ROLES & PERMISSIONS MATRIX
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. Roles & Permissions")
ws3.sheet_properties.tabColor = "7C3AED"

roles_header = ["Module / Feature", "SuperAdmin", "Hospital Admin", "Doctor", "Nurse", "Chemist / Pharmacist", "Patient"]

role_data = [
    roles_header,
    ["— USER MANAGEMENT —", "", "", "", "", "", ""],
    ["Create/Edit Hospital", "Full", "—", "—", "—", "—", "—"],
    ["Create/Edit Users (Doctors, Nurses)", "Full", "Full", "—", "—", "—", "—"],
    ["Assign Roles & Permissions", "Full", "Limited", "—", "—", "—", "—"],
    ["View All Users Across Hospitals", "Full", "—", "—", "—", "—", "—"],
    ["Manage Own Profile", "Full", "Full", "Full", "Full", "Full", "Full"],
    ["", "", "", "", "", "", ""],
    ["— PATIENT MANAGEMENT —", "", "", "", "", "", ""],
    ["Register New Patient", "Full", "Full", "Full", "Full", "—", "—"],
    ["View Patient List", "Full", "Full", "Full", "Full", "Limited", "Own Only"],
    ["Edit Patient Demographics", "Full", "Full", "Full", "Limited", "—", "Own Only"],
    ["View Patient Medical History", "Full", "Full", "Full", "Full", "Limited", "Own Only"],
    ["Manage Emergency Contacts", "Full", "Full", "Full", "Full", "—", "Own Only"],
    ["Upload/View Documents", "Full", "Full", "Full", "Full", "Limited", "Own Only"],
    ["", "", "", "", "", "", ""],
    ["— ONCOLOGY / CANCER MODULES —", "", "", "", "", "", ""],
    ["Create Cancer Diagnosis Record", "—", "—", "Full", "—", "—", "—"],
    ["Edit Diagnosis Details", "Full", "—", "Full", "—", "—", "—"],
    ["View Diagnosis Details", "Full", "Full", "Full", "Full", "View", "Own Only"],
    ["Assign Chemo Protocol", "—", "—", "Full", "—", "—", "—"],
    ["View Treatment Plan", "Full", "Full", "Full", "Full", "Full", "Own Only"],
    ["", "", "", "", "", "", ""],
    ["— CHEMOTHERAPY TRACKING —", "", "", "", "", "", ""],
    ["Schedule Chemo Sessions", "—", "—", "Full", "Full", "—", "—"],
    ["Mark Session Complete", "—", "—", "Full", "Full", "—", "—"],
    ["Log Missed Session", "—", "—", "Full", "Full", "—", "—"],
    ["View Chemo Calendar", "Full", "Full", "Full", "Full", "Full", "Own Only"],
    ["View Remaining Sessions Count", "Full", "Full", "Full", "Full", "Full", "Own Only"],
    ["Pre-Chemo Checklist", "—", "—", "Full", "Full", "—", "—"],
    ["", "", "", "", "", "", ""],
    ["— MEDICATIONS —", "", "", "", "", "", ""],
    ["Prescribe Medications", "—", "—", "Full", "—", "—", "—"],
    ["Verify/Dispense Medications", "—", "—", "—", "—", "Full", "—"],
    ["View Medication List", "Full", "Full", "Full", "Full", "Full", "Own Only"],
    ["Log Extra/Emergency Meds", "—", "—", "Full", "Full", "Full", "—"],
    ["Drug Interaction Alerts", "—", "—", "View", "View", "Full", "—"],
    ["", "", "", "", "", "", ""],
    ["— NOTIFICATIONS —", "", "", "", "", "", ""],
    ["Configure Notification Rules", "Full", "Full", "—", "—", "—", "—"],
    ["Send Manual Notifications", "Full", "Full", "Full", "Full", "—", "—"],
    ["Receive Treatment Alerts", "—", "—", "Full", "Full", "Full", "Full"],
    ["Manage Notification Preferences", "Full", "Full", "Full", "Full", "Full", "Full"],
    ["View Notification History", "Full", "Full", "Own", "Own", "Own", "Own"],
    ["", "", "", "", "", "", ""],
    ["— BENCHMARKS & ANALYTICS —", "", "", "", "", "", ""],
    ["View Hospital-Wide Analytics", "Full", "Full", "—", "—", "—", "—"],
    ["View Patient Benchmarks", "Full", "Full", "Full", "Full", "—", "Own Only"],
    ["View Tumor Marker Trends", "Full", "Full", "Full", "Full", "—", "Own Only"],
    ["Compare Hospital vs National", "Full", "Full", "Full", "—", "—", "—"],
    ["Export Reports", "Full", "Full", "Full", "Limited", "—", "—"],
    ["", "", "", "", "", "", ""],
    ["— ER & EMERGENCY —", "", "", "", "", "", ""],
    ["Log ER Visit", "—", "—", "Full", "Full", "—", "—"],
    ["View ER History", "Full", "Full", "Full", "Full", "—", "Own Only"],
    ["Log Emergency Medications", "—", "—", "Full", "Full", "Full", "—"],
    ["ER Trend Analytics", "Full", "Full", "Full", "—", "—", "—"],
    ["", "", "", "", "", "", ""],
    ["— REPORTS —", "", "", "", "", "", ""],
    ["Generate Patient Reports", "Full", "Full", "Full", "Limited", "—", "Own Only"],
    ["Generate Hospital Reports", "Full", "Full", "—", "—", "—", "—"],
    ["Insurance/Billing Reports", "Full", "Full", "Full", "—", "—", "Own Only"],
    ["Audit Trail Reports", "Full", "Full", "—", "—", "—", "—"],
]

for r, row in enumerate(role_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = PURPLE_FILL
        elif val and val.startswith("—") and val.endswith("—"):
            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        elif val == "Full":
            cell.fill = GREEN_FILL
            cell.font = Font(name="Calibri", bold=True, size=11, color="065F46")
        elif val == "Limited" or val == "View":
            cell.fill = YELLOW_FILL
            cell.font = Font(name="Calibri", size=11, color="92400E")
        elif val == "Own Only" or val == "Own":
            cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
            cell.font = Font(name="Calibri", size=11, color="3730A3")
        elif val == "—":
            cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            cell.font = Font(name="Calibri", size=11, color="9CA3AF")
        else:
            cell.font = BODY_FONT

ws3.column_dimensions["A"].width = 38
for col_letter in ["B", "C", "D", "E", "F", "G"]:
    ws3.column_dimensions[col_letter].width = 22


# ═══════════════════════════════════════════════════════════════════
# TAB 4 — CANCER MODULES DETAIL
# ═══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. Cancer Modules Detail")
ws4.sheet_properties.tabColor = "DC2626"

cancer_data = [
    ["Cancer Type", "Diagnosis Fields", "Chemo Protocols", "Key Tumor Markers", "Benchmark Metrics", "Special Tracking", "Notifications"],
    [
        "Breast Cancer",
        "• Tumor Size & Grade\n• ER/PR Status (+/-)\n• HER2 Status (+/-)\n• BRCA1/BRCA2 mutation\n• Lymph node involvement\n• TNM Staging (I-IV)\n• Histology type",
        "• AC-T (Adriamycin + Cyclophosphamide → Taxol)\n• TC (Taxotere + Cyclophosphamide)\n• TCH (for HER2+)\n• CMF (Cyclophosphamide + Methotrexate + 5-FU)\n• Trastuzumab (Herceptin)\n• Hormone Therapy (Tamoxifen/AI)",
        "• CA 15-3\n• CA 27.29\n• CEA\n• HER2/neu levels",
        "• CA 15-3 trend (↑↓)\n• Tumor size reduction %\n• Chemo completion rate\n• Recurrence-free survival\n• Treatment response (CR/PR/SD/PD)",
        "• Hormone therapy compliance\n• Mammography follow-up\n• Genetic counseling status\n• Reconstruction timeline\n• Lymphedema monitoring",
        "• Missed chemo → Patient + Emergency Contact\n• Rising CA 15-3 → Oncologist alert\n• HER2 therapy due → Nurse reminder\n• Mammography overdue → Patient SMS\n• Hormone therapy refill → Patient + Pharmacy"
    ],
    [
        "Prostate Cancer",
        "• Gleason Score (3+3 to 5+5)\n• PSA Level (ng/mL)\n• TNM Staging\n• Risk Group (Low/Inter/High)\n• Biopsy results\n• Digital rectal exam\n• Bone scan results",
        "• Docetaxel + Prednisone\n• Cabazitaxel\n• Mitoxantrone\n• Abiraterone + Prednisone\n• Enzalutamide\n• ADT (LHRH agonists/antagonists)\n• Radiation combos",
        "• PSA (ng/mL)\n• PSA Velocity\n• PSA Doubling Time\n• Testosterone levels\n• Alkaline Phosphatase",
        "• PSA trend (↑↓) over time\n• PSA response (>50% drop)\n• Time to PSA nadir\n• Chemo completion rate\n• Bone metastasis progression\n• Overall survival benchmarks",
        "• ADT compliance monitoring\n• Bone density scans\n• Cardiovascular risk tracking\n• Sexual function assessment\n• Urinary function tracking",
        "• PSA rising above threshold → Oncologist\n• Missed ADT injection → Patient + Nurse\n• Bone scan due → Scheduling alert\n• Chemo session missed → Patient + Emg Contact\n• Lab work overdue → Patient reminder"
    ],
    [
        "Lung Cancer",
        "• Type: SCLC vs NSCLC\n• Subtype (Adeno/Squamous/Large)\n• Staging (I-IV / Limited-Extensive)\n• Mutations (EGFR, ALK, ROS1, PD-L1%)\n• Pulmonary Function Tests\n• Biopsy/Pathology\n• Performance Status (ECOG)",
        "• Cisplatin + Etoposide (SCLC)\n• Carboplatin + Paclitaxel\n• Carboplatin + Pemetrexed (non-squamous)\n• Pembrolizumab (immunotherapy)\n• Atezolizumab + Chemo\n• Targeted: Osimertinib, Crizotinib\n• Concurrent chemoradiation",
        "• CYFRA 21-1\n• NSE (for SCLC)\n• CEA\n• SCC Antigen\n• PD-L1 expression %",
        "• Tumor marker trends (↑↓)\n• PFT trend (FEV1, DLCO)\n• Treatment response (RECIST)\n• Progression-free survival\n• Immunotherapy response rate\n• ECOG score changes",
        "• Pulmonary function monitoring\n• Oxygen saturation tracking\n• Immunotherapy side effects\n• Brain metastasis screening\n• Smoking cessation status\n• Radiation pneumonitis watch",
        "• PFT declining → Pulmonologist alert\n• Immunotherapy reaction → ER + Oncologist\n• Missed chemo → Patient + Emergency Contact\n• Brain MRI due → Scheduling alert\n• Mutation results ready → Oncologist notification"
    ],
    [
        "Colorectal Cancer",
        "• Tumor Location (Colon/Rectum/Segment)\n• TNM Staging (I-IV)\n• MSI Status (MSI-H/MSS)\n• KRAS/NRAS/BRAF mutation\n• Grade (Well/Mod/Poorly diff)\n• Lymph node count\n• Colonoscopy findings",
        "• FOLFOX (5-FU + Leucovorin + Oxaliplatin)\n• FOLFIRI (5-FU + Leucovorin + Irinotecan)\n• CAPOX (Capecitabine + Oxaliplatin)\n• Bevacizumab (Avastin)\n• Cetuximab (for KRAS wild-type)\n• Pembrolizumab (for MSI-H)\n• Adjuvant vs Palliative regimens",
        "• CEA (ng/mL)\n• CA 19-9\n• LDH\n• Microsatellite status",
        "• CEA trend (↑↓)\n• Treatment response (RECIST)\n• Chemo completion rate\n• Neuropathy grade tracking\n• Surgical margin status\n• 5-year survival benchmarks",
        "• Colonoscopy follow-up scheduling\n• Surgical timeline tracking\n• Oxaliplatin neuropathy grading\n• Stoma care (if applicable)\n• Liver metastasis monitoring\n• Post-surgery recovery milestones",
        "• CEA rising → Oncologist alert\n• Colonoscopy overdue → Patient reminder\n• Neuropathy worsening → Dose adjustment alert\n• Missed chemo → Patient + Emergency Contact\n• Surgery follow-up due → Surgeon + Patient\n• KRAS results ready → Oncologist notification"
    ],
]

for r, row in enumerate(cancer_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL

ws4.column_dimensions["A"].width = 20
for col_letter in ["B", "C", "D", "E", "F", "G"]:
    ws4.column_dimensions[col_letter].width = 42
for r in range(2, len(cancer_data) + 1):
    ws4.row_dimensions[r].height = 200


# ═══════════════════════════════════════════════════════════════════
# TAB 5 — NOTIFICATION SYSTEM
# ═══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5. Notification System")
ws5.sheet_properties.tabColor = "EA580C"

notif_data = [
    ["#", "Trigger Event", "Channel", "Primary Recipient", "Secondary Recipient", "Message Summary", "Frequency", "Priority"],
    # Chemo related
    ["1", "Chemo session missed", "Email + SMS + Push", "Patient", "Emergency Contact + Doctor", "You missed your chemo session on [date]. Please contact your care team to reschedule immediately.", "Immediate + 24hr follow-up", "CRITICAL"],
    ["2", "Chemo session reminder (upcoming)", "Email + SMS + Push", "Patient", "Nurse", "Your next chemotherapy session is scheduled for [date] at [time]. Please arrive 30 min early for blood work.", "48hr + 24hr + 2hr before", "HIGH"],
    ["3", "Chemo cycle completed", "Email + Push", "Patient", "Doctor + Nurse", "Congratulations! You have completed cycle [X] of [Y]. Your next session is on [date].", "After each cycle", "MEDIUM"],
    ["4", "All chemo cycles completed", "Email + SMS", "Patient", "Doctor + Emergency Contact", "Your chemotherapy treatment is complete. Follow-up appointment scheduled for [date].", "Once", "HIGH"],
    # Doctor related
    ["5", "Doctor unavailable / on leave", "Email + SMS", "Patient", "Hospital Admin + Backup Doctor", "Dr. [Name] is unavailable on [date]. Your appointment has been reassigned to Dr. [Backup]. Contact us if you need to reschedule.", "Immediate", "HIGH"],
    ["6", "Doctor assignment change", "Email", "Patient", "New Doctor + Old Doctor", "Your oncologist has been changed to Dr. [Name]. Please contact the clinic for any questions.", "Immediate", "MEDIUM"],
    # Benchmark related
    ["7", "Tumor marker rising (above threshold)", "Email + Push", "Doctor", "Nurse", "ALERT: Patient [Name]'s [Marker] has increased from [X] to [Y]. Review and assess treatment plan.", "Immediate", "CRITICAL"],
    ["8", "Tumor marker improving", "Email + Push", "Patient", "Doctor", "Great news! Your [Marker] levels are trending downward. Keep following your treatment plan.", "After each result", "LOW"],
    ["9", "Benchmark deviation (patient vs expected)", "Email", "Doctor", "Hospital Admin", "Patient [Name]'s treatment response is below expected benchmarks for [Cancer Type] Stage [X].", "Weekly digest", "HIGH"],
    # ER related
    ["10", "ER visit logged", "Email + SMS", "Oncologist", "Nurse + Hospital Admin", "ALERT: Patient [Name] was admitted to ER on [date]. Reason: [Category]. Review medications and treatment plan.", "Immediate", "CRITICAL"],
    ["11", "Extra medication administered", "Email + Push", "Oncologist", "Chemist/Pharmacist", "Additional medication [Drug] was administered to Patient [Name] during ER visit. Review for interactions.", "Immediate", "HIGH"],
    ["12", "Unplanned hospitalization", "Email + SMS", "Doctor + Nurse", "Emergency Contact", "Patient [Name] has been hospitalized unexpectedly. Reason: [Details]. Oncology team has been notified.", "Immediate", "CRITICAL"],
    # Follow-up
    ["13", "Lab work / blood test reminder", "Email + SMS", "Patient", "Nurse", "Your blood work is due before your next chemo session on [date]. Please visit the lab by [deadline].", "72hr + 24hr before", "MEDIUM"],
    ["14", "Follow-up scan due", "Email + SMS + Push", "Patient", "Doctor", "Your follow-up [scan type] is due. Please schedule your appointment before [date].", "1 week + 3 days before deadline", "MEDIUM"],
    ["15", "Treatment plan review due", "Email", "Doctor", "Nurse", "Patient [Name]'s treatment plan is due for review (last reviewed: [date]). Please assess and update.", "Monthly or per protocol", "MEDIUM"],
    # System
    ["16", "Patient registered / onboarded", "Email", "Patient", "Doctor + Nurse", "Welcome to [Hospital] Oncology Care. Your treatment team and portal access details are ready.", "Once", "LOW"],
    ["17", "Insurance expiration warning", "Email + SMS", "Patient", "Hospital Admin", "Your insurance coverage is expiring on [date]. Please update your insurance information.", "30 days + 7 days before", "MEDIUM"],
    ["18", "Emergency contact not set", "Push + Email", "Patient", "Nurse", "Please add an emergency contact to your profile. This is required for your safety during treatment.", "Weekly until resolved", "HIGH"],
]

for r, row in enumerate(notif_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws5.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = ORANGE_FILL
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL
            # Color-code priority
            if c == 8:
                if val == "CRITICAL":
                    cell.fill = RED_FILL
                    cell.font = Font(name="Calibri", bold=True, color="991B1B")
                elif val == "HIGH":
                    cell.fill = YELLOW_FILL
                    cell.font = Font(name="Calibri", bold=True, color="92400E")
                elif val == "MEDIUM":
                    cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
                elif val == "LOW":
                    cell.fill = GREEN_FILL

ws5.column_dimensions["A"].width = 6
ws5.column_dimensions["B"].width = 38
ws5.column_dimensions["C"].width = 22
ws5.column_dimensions["D"].width = 22
ws5.column_dimensions["E"].width = 30
ws5.column_dimensions["F"].width = 65
ws5.column_dimensions["G"].width = 28
ws5.column_dimensions["H"].width = 14


# ═══════════════════════════════════════════════════════════════════
# TAB 6 — BENCHMARKS & ANALYTICS
# ═══════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6. Benchmarks & Analytics")
ws6.sheet_properties.tabColor = "0D9488"

bench_data = [
    ["Dashboard Section", "Metric / KPI", "Visualization", "Data Source", "Update Frequency", "Who Sees It", "Benchmark Comparison"],
    # Patient-level
    ["PATIENT DASHBOARD", "", "", "", "", "", ""],
    ["Treatment Progress", "Chemo cycles completed (X of Y)", "Progress bar + counter", "Chemo schedule module", "After each session", "Patient, Doctor, Nurse", "Expected vs Actual timeline"],
    ["Treatment Progress", "Days since last treatment", "Counter + status badge", "Chemo schedule module", "Real-time", "Patient, Doctor", "Protocol-defined interval"],
    ["Treatment Progress", "Next session countdown", "Calendar card + timer", "Chemo schedule module", "Real-time", "Patient, Nurse", "—"],
    ["Tumor Markers", "Marker value over time (e.g., CA 15-3, PSA, CEA)", "Line chart with trend arrow (↑↓→)", "Lab results integration", "After each lab result", "Patient, Doctor", "Normal range reference lines"],
    ["Tumor Markers", "Marker change % (since last test)", "KPI card with color coding", "Lab results", "After each lab result", "Patient, Doctor, Nurse", "Expected response range"],
    ["Side Effects", "Side effect severity score", "Radar/spider chart", "Patient self-reporting", "Weekly", "Patient, Doctor, Nurse", "Average for same protocol"],
    ["Medications", "Medication adherence %", "Gauge chart", "Prescription module", "Daily", "Patient, Nurse, Chemist", "Target: >95%"],
    ["", "", "", "", "", "", ""],
    ["HOSPITAL DASHBOARD", "", "", "", "", "", ""],
    ["Chemo Operations", "Total active patients per cancer type", "Donut chart by cancer type", "Patient registry", "Real-time", "SuperAdmin, Hospital Admin", "Capacity benchmarks"],
    ["Chemo Operations", "Chemo completion rate %", "Bar chart per cancer type", "Chemo schedule module", "Weekly", "Admin, Doctor", "National average"],
    ["Chemo Operations", "Missed session rate %", "Trend line chart", "Chemo schedule module", "Weekly", "Admin, Doctor, Nurse", "Hospital target (<5%)"],
    ["Chemo Operations", "Average delay between cycles (days)", "Box plot", "Chemo schedule module", "Monthly", "Admin, Doctor", "Protocol-defined intervals"],
    ["Treatment Outcomes", "Treatment response rates (CR/PR/SD/PD)", "Stacked bar per cancer type", "Oncology assessments", "Monthly", "Admin, Doctor", "Published clinical trial rates"],
    ["Treatment Outcomes", "Recurrence rate %", "Trend line per cancer type", "Follow-up records", "Quarterly", "Admin, Doctor", "National cancer database"],
    ["Treatment Outcomes", "Average time to treatment response", "KPI cards per cancer type", "Lab + imaging results", "Monthly", "Admin, Doctor", "Published benchmarks"],
    ["ER & Emergency", "ER visits per 100 chemo cycles", "Bar chart + trend", "ER visit logs", "Monthly", "Admin, Doctor", "National oncology ER benchmark"],
    ["ER & Emergency", "Top ER visit reasons", "Pie chart / treemap", "ER visit logs", "Monthly", "Admin, Doctor, Nurse", "—"],
    ["ER & Emergency", "Unplanned hospitalization rate", "KPI card with trend", "ER visit logs", "Monthly", "Admin", "CMS benchmark"],
    ["ER & Emergency", "Extra medications frequency", "Heat map by drug type", "Medication logs", "Monthly", "Admin, Doctor, Chemist", "—"],
    ["Financial Metrics", "Cost per chemo cycle (avg)", "Bar chart per cancer type", "Billing module", "Monthly", "Admin", "Regional average"],
    ["Financial Metrics", "ER cost impact", "KPI card", "ER + Billing", "Monthly", "Admin", "—"],
    ["Staff Performance", "Patients per oncologist", "Bar chart", "Assignment records", "Real-time", "Admin", "Recommended ratio"],
    ["Staff Performance", "Average treatment plan adherence", "Gauge per doctor", "Protocol compliance", "Monthly", "Admin, Doctor", "Target: >90%"],
]

for r, row in enumerate(bench_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws6.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = TEAL_FILL
        elif val in ("PATIENT DASHBOARD", "HOSPITAL DASHBOARD"):
            cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            for cc in range(2, 8):
                ws6.cell(row=r, column=cc).fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL

ws6.column_dimensions["A"].width = 24
ws6.column_dimensions["B"].width = 42
ws6.column_dimensions["C"].width = 30
ws6.column_dimensions["D"].width = 28
ws6.column_dimensions["E"].width = 20
ws6.column_dimensions["F"].width = 30
ws6.column_dimensions["G"].width = 32


# ═══════════════════════════════════════════════════════════════════
# TAB 7 — ER & EMERGENCY TRACKING
# ═══════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("7. ER & Emergency Tracking")
ws7.sheet_properties.tabColor = "B91C1C"

er_data = [
    ["Category", "Field / Feature", "Description", "Captured By", "Visible To", "Triggers Notification?"],
    ["ER VISIT LOGGING", "", "", "", "", ""],
    ["Visit Details", "Visit Date & Time", "Date/time patient arrived at ER", "Nurse / ER Staff", "Doctor, Nurse, Admin", "Yes — Oncologist notified"],
    ["Visit Details", "ER Facility", "Which hospital/ER location", "Nurse / ER Staff", "Doctor, Admin", "No"],
    ["Visit Details", "Reason Category", "Dropdown: Fever, Infection, Pain, Chemo Reaction, Bleeding, Breathing Difficulty, Fall, Other", "Nurse / ER Staff", "Doctor, Nurse, Admin, Patient", "Yes — if chemo-related"],
    ["Visit Details", "Detailed Description", "Free-text clinical notes about the visit", "Doctor / Nurse", "Doctor, Nurse", "No"],
    ["Visit Details", "Severity Level", "Triage: Critical / Urgent / Standard", "ER Staff", "All Care Team", "Yes — if Critical"],
    ["Visit Details", "Outcome", "Treated & Released / Admitted / Transferred / Against Medical Advice", "Doctor", "All Care Team + Admin", "Yes — if Admitted"],
    ["Visit Details", "Duration of Stay", "Hours in ER / Days if hospitalized", "System (calculated)", "Doctor, Admin", "No"],
    ["", "", "", "", "", ""],
    ["EMERGENCY MEDICATIONS", "", "", "", "", ""],
    ["Extra Meds", "Medication Name", "Drug administered during ER visit", "Doctor / Nurse", "Doctor, Nurse, Chemist", "Yes — Pharmacist alert"],
    ["Extra Meds", "Dosage & Route", "Amount + IV/Oral/IM/SC", "Doctor / Nurse", "Doctor, Nurse, Chemist", "No"],
    ["Extra Meds", "Reason for Extra Med", "Why this med was needed (linked to ER reason)", "Doctor", "Doctor, Chemist", "No"],
    ["Extra Meds", "Drug Interaction Check", "Auto-check against current chemo + support meds", "System (auto)", "Doctor, Chemist", "Yes — if interaction found"],
    ["Extra Meds", "Impact on Treatment Plan", "Chemo delayed / Protocol modified / No change", "Doctor", "Doctor, Nurse, Admin", "Yes — if plan changed"],
    ["", "", "", "", "", ""],
    ["FOLLOW-UP ACTIONS", "", "", "", "", ""],
    ["Follow-up", "Post-ER Assessment Required", "Flag: Does oncologist need to reassess?", "ER Doctor", "Oncologist", "Yes"],
    ["Follow-up", "Chemo Schedule Impact", "Next session: On time / Delayed / Cancelled", "Oncologist", "Patient, Nurse", "Yes — Patient notified"],
    ["Follow-up", "Additional Tests Ordered", "Labs, imaging, or consults ordered post-ER", "Doctor", "Nurse, Patient", "Yes — scheduling"],
    ["Follow-up", "Emergency Contact Notified", "Was emergency contact called? When?", "Nurse", "Admin", "Auto-logged"],
    ["", "", "", "", "", ""],
    ["ANALYTICS (Dashboard)", "", "", "", "", ""],
    ["Analytics", "ER visits per patient (over time)", "Track frequency of ER visits per patient", "System", "Doctor, Admin", "—"],
    ["Analytics", "ER visits by cancer type", "Compare ER rates across Breast/Prostate/Lung/Colorectal", "System", "Admin", "—"],
    ["Analytics", "Most common ER reasons", "Top reasons by category", "System", "Admin, Doctor", "—"],
    ["Analytics", "ER visits post-chemo (within 7 days)", "How many ER visits occur shortly after chemo", "System", "Admin, Doctor", "—"],
    ["Analytics", "Extra medication frequency", "Most commonly administered emergency drugs", "System", "Admin, Chemist", "—"],
    ["Analytics", "Cost per ER episode", "Average cost of ER visits per cancer type", "Billing", "Admin", "—"],
]

for r, row in enumerate(er_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws7.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="B91C1C", end_color="B91C1C", fill_type="solid")
        elif val in ("ER VISIT LOGGING", "EMERGENCY MEDICATIONS", "FOLLOW-UP ACTIONS", "ANALYTICS (Dashboard)"):
            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
            for cc in range(2, 7):
                ws7.cell(row=r, column=cc).fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL

ws7.column_dimensions["A"].width = 24
ws7.column_dimensions["B"].width = 35
ws7.column_dimensions["C"].width = 55
ws7.column_dimensions["D"].width = 22
ws7.column_dimensions["E"].width = 30
ws7.column_dimensions["F"].width = 25


# ═══════════════════════════════════════════════════════════════════
# TAB 8 — TECH STACK & ARCHITECTURE
# ═══════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("8. Tech Stack")
ws8.sheet_properties.tabColor = "6366F1"

tech_data = [
    ["Layer", "Technology", "Purpose", "Notes"],
    ["Frontend", "React 18 + TypeScript", "Patient portal & Hospital dashboard SPA", "Vite build tool, responsive design"],
    ["Frontend", "Tailwind CSS / Ant Design", "UI component library & styling", "Consistent medical-grade UI"],
    ["Frontend", "React Query (TanStack)", "API state management & caching", "Optimistic updates, real-time sync"],
    ["Frontend", "Chart.js / Recharts", "Benchmark visualizations & trend charts", "Tumor marker graphs, analytics"],
    ["Frontend", "React Router v6", "Client-side routing & role-based routes", "Protected routes per role"],
    ["", "", "", ""],
    ["Backend", "NestJS + TypeScript", "REST API server & business logic", "Modular architecture, guards, interceptors"],
    ["Backend", "TypeORM", "Database ORM & migrations", "Entity-based models, relations"],
    ["Backend", "PostgreSQL", "Primary relational database", "HIPAA-compliant data storage"],
    ["Backend", "Redis", "Caching, session store, rate limiting", "Notification queue backing"],
    ["Backend", "JWT + Passport", "Authentication & authorization", "Role-based guards, refresh tokens"],
    ["", "", "", ""],
    ["Notifications", "SendGrid / AWS SES", "Email notification delivery", "Templates, tracking, compliance"],
    ["Notifications", "Twilio", "SMS notification delivery", "Patient reminders, emergency alerts"],
    ["Notifications", "Firebase Cloud Messaging", "Push notifications", "Mobile + browser push"],
    ["Notifications", "Bull Queue (Redis)", "Async notification processing", "Retry logic, scheduled delivery"],
    ["", "", "", ""],
    ["Infrastructure", "Docker + Docker Compose", "Containerization", "Dev/staging/production parity"],
    ["Infrastructure", "AWS / GCP", "Cloud hosting", "HIPAA-eligible services"],
    ["Infrastructure", "GitHub Actions", "CI/CD pipeline", "Automated testing + deployment"],
    ["Infrastructure", "Nginx", "Reverse proxy & load balancer", "SSL termination, rate limiting"],
    ["", "", "", ""],
    ["Security", "HIPAA Compliance", "Healthcare data protection", "Encryption at rest + transit, audit logs"],
    ["Security", "RBAC (Role-Based Access)", "Permission enforcement", "6 roles with granular permissions"],
    ["Security", "Data Encryption", "AES-256 at rest, TLS 1.3 in transit", "PHI protection"],
    ["Security", "Audit Logging", "Track all data access & changes", "Compliance reporting"],
    ["", "", "", ""],
    ["Testing", "Jest", "Unit & integration testing", "Backend + Frontend"],
    ["Testing", "Cypress / Playwright", "End-to-end testing", "Critical workflow coverage"],
    ["Testing", "k6 / Artillery", "Load & performance testing", "Chemo scheduling stress tests"],
]

for r, row in enumerate(tech_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws8.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL

ws8.column_dimensions["A"].width = 20
ws8.column_dimensions["B"].width = 32
ws8.column_dimensions["C"].width = 45
ws8.column_dimensions["D"].width = 45


# ═══════════════════════════════════════════════════════════════════
# TAB 9 — MILESTONES & DELIVERABLES
# ═══════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("9. Milestones & Deliverables")
ws9.sheet_properties.tabColor = "059669"

milestone_data = [
    ["#", "Milestone", "Target Date", "Deliverables", "Acceptance Criteria", "Dependencies", "Status"],
    ["M1", "Project Kickoff & Setup", "Week 2",
     "• Monorepo configured (React + NestJS)\n• DB schema v1 finalized\n• CI/CD pipeline running\n• Dev environment ready",
     "• All team members can run project locally\n• CI builds pass on main branch\n• Schema reviewed and approved",
     "Client approval on requirements", "Not Started"],
    ["M2", "Auth & Role System Complete", "Week 4",
     "• Login/Register/Forgot Password\n• All 6 roles implemented\n• Permission matrix enforced\n• Audit logging active",
     "• Each role can only access permitted routes\n• JWT refresh working\n• Admin can CRUD users",
     "M1 complete", "Not Started"],
    ["M3", "Patient Module & Breast Cancer", "Week 8",
     "• Patient registration flow\n• Breast cancer diagnosis & chemo tracking\n• First cancer module fully functional\n• Patient portal v1",
     "• Doctor can create diagnosis + assign chemo\n• Patient sees treatment progress\n• Tumor markers tracked",
     "M2 complete", "Not Started"],
    ["M4", "All 4 Cancer Modules Complete", "Week 14",
     "• Prostate, Lung, Colorectal modules\n• All chemo protocols implemented\n• Tumor marker tracking for each\n• Cancer-specific benchmarks",
     "• All 4 cancer types fully functional\n• Protocol-specific workflows working\n• Marker trend charts displaying",
     "M3 complete", "Not Started"],
    ["M5", "Chemo Engine & Medications", "Week 16",
     "• Chemo scheduling calendar\n• Missed session detection\n• Medication management\n• Chemist verification flow",
     "• Sessions auto-schedule per protocol\n• Missed sessions flagged within 24hr\n• Drug interactions detected",
     "M4 complete", "Not Started"],
    ["M6", "Notification System Live", "Week 18",
     "• Email + SMS + Push working\n• All 18 notification triggers active\n• Alert rules engine running\n• User preference management",
     "• Missed chemo triggers notification\n• Doctor unavailable alerts patient\n• ER visit notifies oncologist",
     "M5 complete, SendGrid/Twilio setup", "Not Started"],
    ["M7", "Benchmarks & Analytics Dashboard", "Week 20",
     "• Patient + Hospital dashboards\n• Tumor marker trend charts\n• ER analytics\n• Treatment outcome benchmarks",
     "• All chart types rendering correctly\n• Data refreshes per defined frequency\n• Role-based visibility enforced",
     "M4 + M5 complete", "Not Started"],
    ["M8", "ER & Emergency Module", "Week 20",
     "• ER visit logging\n• Emergency medication tracking\n• Drug interaction checker\n• ER analytics dashboard",
     "• ER visits logged with all required fields\n• Oncologist auto-notified\n• Extra meds checked for interactions",
     "M6 complete", "Not Started"],
    ["M9", "Reports & Compliance", "Week 24",
     "• PDF report generation\n• Insurance/billing exports\n• Audit trail reports\n• Custom report builder",
     "• Reports export correctly as PDF\n• Audit logs complete and searchable\n• HIPAA compliance checklist passed",
     "M7 complete", "Not Started"],
    ["M10", "UAT & Production Launch", "Week 28",
     "• Full QA testing complete\n• Security audit passed\n• UAT with hospital staff\n• Production deployment\n• Training materials delivered",
     "• >80% test coverage\n• Zero critical bugs\n• Hospital staff trained\n• System stable under load\n• Documentation complete",
     "M1–M9 complete", "Not Started"],
]

for r, row in enumerate(milestone_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws9.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL

ws9.column_dimensions["A"].width = 8
ws9.column_dimensions["B"].width = 34
ws9.column_dimensions["C"].width = 14
ws9.column_dimensions["D"].width = 45
ws9.column_dimensions["E"].width = 45
ws9.column_dimensions["F"].width = 30
ws9.column_dimensions["G"].width = 14
for r in range(2, len(milestone_data) + 1):
    ws9.row_dimensions[r].height = 90


# ═══════════════════════════════════════════════════════════════════
# TAB 10 — PATIENT PORTAL SCREENS
# ═══════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("10. Patient Portal Screens")
ws10.sheet_properties.tabColor = "8B5CF6"

portal_data = [
    ["Screen / Page", "Key Components", "Data Displayed", "Actions Available", "Notifications Shown"],
    ["Dashboard (Home)", "• Treatment progress card\n• Next appointment card\n• Tumor marker mini-chart\n• Recent notifications\n• Quick actions",
     "• Chemo cycles: X of Y completed\n• Days until next session\n• Latest marker value with trend\n• Upcoming lab work dates",
     "• View full schedule\n• Contact care team\n• Report symptoms\n• View notifications",
     "• Upcoming appointment reminders\n• Missed session alerts\n• New lab results available"],
    ["My Treatment Plan", "• Cancer type & stage summary\n• Chemo protocol details\n• Cycle-by-cycle timeline\n• Medication list",
     "• Full treatment protocol\n• Each cycle: date, drugs, status\n• Completed ✓ / Upcoming / Missed ✗\n• Expected completion date",
     "• Download treatment summary\n• View cycle details\n• See medication info",
     "• Upcoming cycle reminders\n• Protocol changes by doctor"],
    ["Chemo Schedule", "• Calendar view (month/week)\n• Session detail cards\n• Pre-chemo checklist status\n• History log",
     "• All scheduled sessions\n• Session times & locations\n• Required pre-visit labs\n• Completed session summaries",
     "• View session details\n• See pre-chemo requirements\n• Download schedule",
     "• 48hr + 24hr + 2hr reminders\n• Lab work due alerts"],
    ["My Health Metrics", "• Tumor marker trend chart\n• Side effect tracker\n• Weight/vitals log\n• Treatment response indicator",
     "• Marker values over time (line chart)\n• Trend arrows (↑↓→)\n• Normal range reference\n• Side effect severity scores",
     "• Log symptoms\n• Log weight/vitals\n• View historical data\n• Download reports",
     "• Marker improving/worsening alerts\n• Time for next lab work"],
    ["My Medications", "• Current medication list\n• Dosage & schedule\n• Refill status\n• Side effect info",
     "• Chemo drugs (per cycle)\n• Supportive medications\n• Hormone therapy (if applicable)\n• Last dispensed dates",
     "• View drug information\n• Request refill\n• Report side effects",
     "• Refill reminders\n• New prescription alerts"],
    ["Messages & Communication", "• Inbox (care team messages)\n• Compose new message\n• Contact directory\n• Urgent contact info",
     "• Messages from doctor/nurse\n• Automated system messages\n• Read/unread status",
     "• Reply to messages\n• New message to care team\n• Emergency call button",
     "• New message notifications\n• Urgent alerts from care team"],
    ["My Documents", "• Lab reports\n• Imaging reports\n• Prescriptions\n• Treatment summaries\n• Insurance documents",
     "• Document list with dates\n• Document type tags\n• Preview thumbnails",
     "• View/download documents\n• Upload documents\n• Share with care team",
     "• New document uploaded by hospital\n• Report ready notification"],
    ["My Profile & Settings", "• Personal information\n• Emergency contacts\n• Insurance details\n• Notification preferences\n• Password/security",
     "• Demographics\n• Emergency contact list\n• Insurance status & expiry\n• Notification channel preferences",
     "• Edit profile\n• Add/edit emergency contacts\n• Update insurance\n• Toggle notification channels\n• Change password",
     "• Missing emergency contact alert\n• Insurance expiration warning\n• Profile incomplete notice"],
]

for r, row in enumerate(portal_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws10.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL

ws10.column_dimensions["A"].width = 26
ws10.column_dimensions["B"].width = 38
ws10.column_dimensions["C"].width = 40
ws10.column_dimensions["D"].width = 35
ws10.column_dimensions["E"].width = 35
for r in range(2, len(portal_data) + 1):
    ws10.row_dimensions[r].height = 130


# ═══════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════
output_path = "/home/tec/tpai/ehr_system_react_nest/Oncology_EHR_Project_Workplan.xlsx"
wb.save(output_path)
print(f"\n✅ Workplan generated successfully!")
print(f"📄 File: {output_path}")
print(f"\n📊 Tabs created:")
print(f"   1. Project Overview")
print(f"   2. Phase-Wise Plan (8 phases, 14 sprints, 28 weeks)")
print(f"   3. Roles & Permissions (6 roles, 55+ permissions)")
print(f"   4. Cancer Modules Detail (Breast, Prostate, Lung, Colorectal)")
print(f"   5. Notification System (18 notification triggers)")
print(f"   6. Benchmarks & Analytics (25+ KPIs)")
print(f"   7. ER & Emergency Tracking (visit logging, extra meds, analytics)")
print(f"   8. Tech Stack (React + NestJS + PostgreSQL)")
print(f"   9. Milestones & Deliverables (10 milestones)")
print(f"  10. Patient Portal Screens (8 screens)")
print(f"\n📌 To use in Google Sheets:")
print(f"   1. Go to Google Sheets → File → Import → Upload")
print(f"   2. Select this .xlsx file")
print(f"   3. Choose 'Replace spreadsheet'")
print(f"   4. All 10 tabs will be imported with formatting")
