#!/usr/bin/env python3
"""
Generate Oncology EHR Project Workplan Excel Workbook — v2
Includes: Basic Infrastructure First + AI Integration Plan
Upload this .xlsx to Google Sheets → File → Import → Upload
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

# ── Shared Styles ─────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
SUB_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="1F2937")
BODY_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="FFFFFF")

DARK_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
GREEN_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
TEAL_FILL = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
AI_FILL = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")  # Indigo for AI
INFRA_FILL = PatternFill(start_color="059669", end_color="059669", fill_type="solid")  # Emerald for Infra

WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def style_header_row(ws, row, num_cols, fill=HEADER_FILL):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_body(ws, start_row, end_row, num_cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if r % 2 == 0:
                cell.fill = LIGHT_FILL


def auto_width(ws, num_cols, min_w=15, max_w=50):
    for col in range(1, num_cols + 1):
        max_len = min_w
        letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = max_len + 2


# ═══════════════════════════════════════════════════════════════════
# TAB 1 — PROJECT OVERVIEW
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Project Overview"
ws1.sheet_properties.tabColor = "1E3A5F"

overview_data = [
    ["Section", "Details"],
    ["Project Name", "Oncology Care Management EHR Platform"],
    ["Project Type", "SaaS — Hospital & Patient-Facing EHR System for Oncology"],
    ["Prepared For", "[Client / Hospital Name]"],
    ["Prepared By", "TPAI Engineering Team"],
    ["Date", "March 2026"],
    ["Version", "2.0 — Infrastructure-First + AI Integration"],
    ["", ""],
    ["VISION", ""],
    ["Purpose",
     "Build a comprehensive, role-based oncology EHR platform that hospitals can deploy to manage cancer patients through their entire treatment journey — from diagnosis through chemotherapy cycles, emergency events, medication management, and ongoing benchmarking. Includes AI-powered document analysis and treatment recommendations."],
    ["Patient Value",
     "Patients get a personal portal to track remaining chemo sessions, view treatment progress, receive notifications for missed appointments, and stay connected with their care team."],
    ["Hospital Value",
     "Hospitals get a powerful dashboard with real-time benchmarks, compliance tracking, ER visit monitoring, role-based workflows for doctors/nurses/chemists, automated notification systems, and AI-powered report analysis."],
    ["AI Value",
     "AI analyzes uploaded medical reports (PDF/scanned documents) using OCR + LLM to extract structured data, auto-create oncology records, calculate AJCC staging, and generate treatment roadmaps based on NCCN guidelines."],
    ["", ""],
    ["TECH STACK", ""],
    ["Frontend", "React 18 + TypeScript + Vite + Tailwind CSS / Ant Design"],
    ["Backend", "NestJS + TypeScript + TypeORM + PostgreSQL"],
    ["AI/ML", "Groq API (Llama 3.3 70B) + Tesseract OCR + pdfplumber + pdf2image"],
    ["Infrastructure", "Docker + Docker Compose + GitHub Actions + Nginx"],
    ["", ""],
    ["CANCER TYPES COVERED", ""],
    ["1", "Breast Cancer — Chemotherapy protocols, hormone therapy tracking, HER2 status monitoring"],
    ["2", "Prostate Cancer — PSA benchmarking, radiation + chemo cycles, hormone therapy tracking"],
    ["3", "Lung Cancer — Staging (SCLC/NSCLC), immunotherapy + chemo tracking, PFT benchmarks"],
    ["4", "Colorectal Cancer — CEA marker tracking, surgical + chemo timelines, colonoscopy scheduling"],
    ["", ""],
    ["KEY CAPABILITIES", ""],
    ["AI Report Analysis", "Upload medical reports (PDF/images) → OCR text extraction → LLM-powered data extraction → Auto-create oncology records"],
    ["AI Treatment Roadmap", "AI generates treatment recommendations (surgery, chemo, radiation, immunotherapy) based on cancer type, stage, biomarkers & NCCN guidelines"],
    ["AJCC Staging", "Deterministic AJCC 8th Edition cancer staging engine for Breast, Prostate, Lung, Colorectal cancers"],
    ["Chemo Tracking", "Track total cycles prescribed vs completed, next session dates, missed sessions with reasons"],
    ["Smart Notifications", "Automated alerts via Email/SMS/Push to patient, emergency contact, and care team"],
    ["Benchmarking", "Tumor marker trends (up/down), treatment response rates, survival benchmarks per cancer type"],
    ["ER Monitoring", "Track emergency room visits, extra medications administered, unplanned hospitalizations"],
    ["Role-Based Access", "SuperAdmin, Hospital Admin, Doctor, Nurse, Chemist/Pharmacist, Patient — each with specific permissions"],
    ["Reporting", "Exportable reports for compliance, insurance, and clinical audits"],
]

for r, row in enumerate(overview_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = DARK_FILL
        elif val in ("VISION", "CANCER TYPES COVERED", "KEY CAPABILITIES", "TECH STACK"):
            cell.font = Font(name="Calibri", bold=True, size=12, color="1E3A5F")
            cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            ws1.cell(row=r, column=2).fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        else:
            cell.font = BODY_FONT

ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 90


# ═══════════════════════════════════════════════════════════════════
# TAB 2 — PHASE-WISE PROJECT PLAN (Infrastructure First + AI)
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Phase-Wise Plan")
ws2.sheet_properties.tabColor = "2563EB"

phases = [
    ["Phase", "Sprint", "Duration", "Module / Deliverable", "Key Tasks", "Owner", "Status", "Priority"],

    # ──────────────────────────────────────────────────────────────
    # PHASE 1: BASIC INFRASTRUCTURE (MUST DO FIRST)
    # ──────────────────────────────────────────────────────────────
    ["Phase 1: Basic Infrastructure", "Sprint 1", "Week 1–2",
     "Environment & Configuration Setup",
     "• Create .env files (Dev/Staging/Prod) with all config variables\n• Configure database connection with environment variables\n• Setup TypeORM migrations (disable synchronize:true)\n• Generate initial migration from existing entities\n• Configure CORS properly for all environments\n• Setup logging (Winston/Pino logger)\n• Health check endpoint (/api/health)",
     "DevOps + Backend Lead", "Not Started", "P0 — Critical"],

    ["Phase 1: Basic Infrastructure", "Sprint 1", "Week 1–2",
     "Docker & Containerization",
     "• Dockerfile for NestJS backend (multi-stage build)\n• Dockerfile for React frontend\n• docker-compose.yml (backend + frontend + PostgreSQL + Redis)\n• docker-compose.dev.yml for local development\n• Volume mounts for hot-reload in dev\n• .dockerignore files\n• Docker health checks",
     "DevOps", "Not Started", "P0 — Critical"],

    ["Phase 1: Basic Infrastructure", "Sprint 1", "Week 1–2",
     "CI/CD Pipeline",
     "• GitHub Actions workflow for backend (lint + test + build)\n• GitHub Actions workflow for frontend (lint + test + build)\n• Automated testing on PR\n• Docker image build & push to registry\n• Dev/Staging deployment automation\n• Branch protection rules\n• Environment secrets management",
     "DevOps", "Not Started", "P0 — Critical"],

    ["Phase 1: Basic Infrastructure", "Sprint 2", "Week 3–4",
     "DTO Validation Layer (Backend)",
     "• Create DTOs for ALL existing endpoints (currently using 'any' types)\n• Patient DTOs: CreatePatientDto, UpdatePatientDto, PatientQueryDto\n• Allergy DTOs: CreateAllergyDto, UpdateAllergyDto\n• Vitals DTOs: CreateVitalsDto, UpdateVitalsDto\n• Medication DTOs: CreateMedicationDto, UpdateMedicationDto\n• ChronicCondition DTOs: CreateChronicConditionDto, UpdateChronicConditionDto\n• Lifestyle/MedicalHistory DTOs\n• Oncology DTOs: CreateOncologyRecordDto, UpdateOncologyRecordDto\n• Treatment, FollowUp, Symptom, Payer DTOs\n• Auth DTOs: LoginDto, RefreshDto\n• Apply class-validator decorators (@IsString, @IsEmail, @IsOptional, etc.)\n• Swagger @ApiProperty decorators on all DTOs",
     "Backend Team", "Not Started", "P0 — Critical"],

    ["Phase 1: Basic Infrastructure", "Sprint 2", "Week 3–4",
     "Error Handling & Response Standardization",
     "• Global exception filter (HttpExceptionFilter)\n• Standardized API response format: { success, data, message, errors }\n• Custom exception classes (PatientNotFoundException, etc.)\n• Validation error formatting (field-level errors)\n• Database error handling (unique constraint, FK violations)\n• Request/Response logging interceptor\n• API versioning strategy (v1 prefix)",
     "Backend Lead", "Not Started", "P0 — Critical"],

    ["Phase 1: Basic Infrastructure", "Sprint 2", "Week 3–4",
     "File Upload Infrastructure",
     "• Multer configuration for file uploads in NestJS\n• File storage service (local disk + S3-compatible)\n• File type validation (PDF, PNG, JPG, JPEG, DICOM)\n• File size limits (configurable per type, e.g., 50MB for PDFs)\n• Secure file serving endpoint with auth\n• File cleanup service (orphaned files)\n• Media directory structure (/uploads/reports/, /uploads/documents/)",
     "Backend Team", "Not Started", "P0 — Critical"],

    # ──────────────────────────────────────────────────────────────
    # PHASE 2: AUTHENTICATION & ROLE-BASED ACCESS
    # ──────────────────────────────────────────────────────────────
    ["Phase 2: Auth & RBAC", "Sprint 3", "Week 5–6",
     "Enhanced Authentication",
     "• Role-based access control (RBAC) with guards\n• Roles: SuperAdmin, Hospital Admin, Doctor, Nurse, Chemist, Patient\n• @Roles() decorator + RolesGuard\n• Register endpoint (POST /api/auth/register)\n• Forgot Password flow (email-based reset token)\n• Change Password endpoint\n• Session management & token blacklisting on logout\n• Refresh token rotation\n• Audit logging for ALL auth events (login, logout, failed attempts)\n• Rate limiting on auth endpoints (throttler)",
     "Backend Lead", "Not Started", "P0 — Critical"],

    ["Phase 2: Auth & RBAC", "Sprint 3", "Week 5–6",
     "User Management Module",
     "• User entity enhancement (role, hospital_id, profile fields)\n• SuperAdmin: CRUD for hospitals and admins\n• Hospital Admin: CRUD for doctors, nurses, chemists\n• Profile management for all roles\n• Role permission matrix implementation\n• User search & filtering\n• User activation/deactivation\n• Swagger docs for all user endpoints",
     "Full Stack Team", "Not Started", "P0 — Critical"],

    # ──────────────────────────────────────────────────────────────
    # PHASE 3: CORE PATIENT & ONCOLOGY (already partially built)
    # ──────────────────────────────────────────────────────────────
    ["Phase 3: Core Patient & Oncology", "Sprint 4", "Week 7–8",
     "Patient Module Hardening",
     "• Apply DTOs to all existing patient endpoints\n• Patient demographics & medical history validation\n• Insurance & emergency contact management\n• Cancer diagnosis intake form\n• Patient search & advanced filtering (date ranges, cancer type, status)\n• Pagination with proper metadata (total, page, limit)\n• Patient photo upload\n• Soft-delete & restore functionality testing",
     "Backend + Frontend", "Not Started", "P1 — High"],

    ["Phase 3: Core Patient & Oncology", "Sprint 5", "Week 9–10",
     "Breast Cancer Module",
     "• Diagnosis details (ER/PR/HER2 status, tumor grade)\n• Chemotherapy protocol assignment (AC-T, TC, etc.)\n• Cycle tracking (prescribed vs completed)\n• Hormone therapy tracking\n• Tumor marker benchmarks (CA 15-3)\n• Treatment response assessment",
     "Oncology Team", "Not Started", "P1 — High"],

    ["Phase 3: Core Patient & Oncology", "Sprint 6", "Week 11–12",
     "Prostate Cancer Module",
     "• Diagnosis (Gleason score, PSA levels, staging)\n• Chemo protocol assignment (Docetaxel, Cabazitaxel)\n• PSA trend benchmarking (rising/falling)\n• Hormone/ADT therapy tracking\n• Radiation therapy integration\n• Treatment response metrics",
     "Oncology Team", "Not Started", "P1 — High"],

    ["Phase 3: Core Patient & Oncology", "Sprint 7", "Week 13–14",
     "Lung Cancer Module",
     "• Diagnosis (SCLC vs NSCLC, staging, mutations: EGFR, ALK, ROS1, PD-L1%)\n• Chemo protocol (Cisplatin/Carboplatin combos)\n• Immunotherapy tracking (Pembrolizumab, etc.)\n• PFT (Pulmonary Function Test) benchmarks\n• Tumor marker monitoring (CYFRA 21-1, NSE)\n• Treatment line tracking (1st, 2nd, 3rd line)",
     "Oncology Team", "Not Started", "P1 — High"],

    ["Phase 3: Core Patient & Oncology", "Sprint 8", "Week 15–16",
     "Colorectal Cancer Module",
     "• Diagnosis (staging, tumor location, MSI status)\n• Chemo protocol (FOLFOX, FOLFIRI, CAPOX)\n• CEA marker trend tracking\n• KRAS/NRAS/BRAF mutation tracking\n• Surgical timeline integration\n• Colonoscopy scheduling & follow-up\n• Treatment response (RECIST criteria)",
     "Oncology Team", "Not Started", "P1 — High"],

    # ──────────────────────────────────────────────────────────────
    # PHASE 4: AI/ML INTEGRATION (Port from Django)
    # ──────────────────────────────────────────────────────────────
    ["Phase 4: AI Integration — OCR & Text Extraction", "Sprint 9", "Week 17–18",
     "MedicalReport Entity & File Upload Pipeline",
     "• Create MedicalReport entity (port from Django model)\n  — Fields: file, document_type, extracted_text, extracted_data (JSONB),\n    insights (JSONB), recommendations (JSONB), status, page_count,\n    estimated_tokens, is_scanned, retry_count\n• Create ProcessingProgress entity (tracks document processing stages)\n• MedicalReport CRUD endpoints\n• File upload endpoint: POST /api/medical-reports/upload\n• Processing status endpoint: GET /api/medical-reports/:id/status\n• Insights endpoint: GET /api/medical-reports/:id/insights\n• Approve endpoint: POST /api/medical-reports/:id/approve\n• Retry endpoint: POST /api/medical-reports/:id/retry\n• DTOs + Swagger docs for all report endpoints",
     "Backend Team", "Not Started", "P1 — High"],

    ["Phase 4: AI Integration — OCR & Text Extraction", "Sprint 9", "Week 17–18",
     "Text Extraction Service (OCR + PDF)",
     "• Install & configure system dependencies (Tesseract, Poppler)\n• Port TextExtractionService from Django → NestJS service\n• PDF text extraction using pdfplumber (Node equivalent: pdf-parse or pdf.js)\n• Scanned PDF detection (< 50 chars = image-based)\n• Scanned PDF OCR: pdf-to-image conversion + Tesseract\n• Image OCR with preprocessing: grayscale → contrast → sharpen → upscale\n• Tesseract config: --oem 3 --psm 3\n• OCR text cleanup: whitespace fixes, artifact removal (|→I),\n  medical term corrections (Gleason, Adenocarcinoma, Pathology Report)\n• Smart routing: try digital extraction first, fallback to OCR\n• Support formats: PDF, PNG, JPG, JPEG, TIFF\n• Unit tests for extraction service",
     "AI/Backend Team", "Not Started", "P1 — High"],

    ["Phase 4: AI Integration — LLM Analysis", "Sprint 10", "Week 19–20",
     "AI Analysis Service (Groq/LLM Integration)",
     "• Install Groq SDK for Node.js (groq-sdk npm package)\n• Configure Groq API: GROQ_API_KEY, model (llama-3.3-70b-versatile),\n  timeout (30s), max_tokens (1500), temperature (0.1)\n• Port AIAnalysisService from Django → NestJS:\n  — Direct mode: full text → single Groq API call → parse JSON response\n  — Chunked mode: text > 6000 tokens → split into overlapping chunks\n    → per-chunk API call → merge results\n• Document-type-specific LLM prompts:\n  — Generic medical report extraction prompt\n  — Lab report prompt (test names, values, units, reference ranges)\n  — Radiology/imaging prompt (findings, measurements, severity)\n  — Prescription prompt (drug names, dosages, frequencies)\n  — Clinical note prompt (diagnoses, procedures, medications)\n  — Pathology report prompt (FULL oncology schema: cancer type,\n    TNM staging, histology, biomarkers, ER/PR/HER2, Gleason,\n    EGFR, ALK, PD-L1, KRAS, BRAF, MSI/MMR, IHC, specimen info)\n• JSON response parsing with validation\n• Error handling: API timeouts, rate limits, malformed responses\n• Retry logic with exponential backoff\n• Unit + integration tests",
     "AI/Backend Team", "Not Started", "P1 — High"],

    ["Phase 4: AI Integration — LLM Analysis", "Sprint 10", "Week 19–20",
     "Document Chunking Service",
     "• Port ChunkingService from Django → NestJS\n• Token estimation (character-count / 4 approximation)\n• Text splitting: 6000 token chunks with 200-token overlap\n• Smart chunk boundaries (avoid splitting mid-sentence)\n• Result merging across chunks:\n  — Oncology data merging (first non-null wins for scalar fields)\n  — Pathology info merging\n  — Vitals, lab results, medications deduplication\n  — Allergies, diagnoses, insights list union\n  — Biomarker/IHC dictionary merging\n• Unit tests for chunking & merging logic",
     "AI/Backend Team", "Not Started", "P1 — High"],

    ["Phase 4: AI Integration — Post-Processing", "Sprint 11", "Week 21–22",
     "Abnormal Lab Detection & Report Type Classification",
     "• Port abnormal lab detection logic to NestJS\n  — Parse reference ranges from AI-extracted lab results\n  — Flag values outside normal ranges (is_abnormal boolean)\n  — Handle edge cases: '<', '>', ranges like '10-20'\n• Port report type auto-detection (NLP keyword matching)\n  — Keyword-based classification: pathology, ultrasound, X-ray,\n    CT scan, MRI, PET scan, mammography, radiology, lab report\n  — Minimum 2 keyword matches for pathology (OCR robustness)\n  — No LLM API call — pure regex/keyword matching\n• Unit tests for both services",
     "AI/Backend Team", "Not Started", "P1 — High"],

    ["Phase 4: AI Integration — AJCC Staging", "Sprint 11", "Week 21–22",
     "AJCC 8th Edition Staging Engine",
     "• Port AJCCStagingEngine from Django → NestJS (deterministic, no AI)\n• Breast cancer staging:\n  — Anatomic staging (T + N + M)\n  — Prognostic staging (T + N + M + Grade + ER/PR/HER2)\n  — Molecular subtype classification (Luminal A/B, HER2+, Triple-negative)\n• Prostate cancer staging:\n  — NCCN Risk Groups (T stage + PSA + Gleason primary/secondary score)\n• Lung cancer staging (NSCLC):\n  — TNM stage groups mapping\n• Colorectal cancer staging:\n  — TNM stage groups mapping\n• Stage-specific treatment recommendations:\n  — Stage IV → URGENT systemic therapy, palliative care, molecular testing\n  — Stage III → Tumor board, neoadjuvant therapy, staging imaging\n  — Stage II → Surgical consultation, adjuvant therapy\n  — Cancer-specific: BRCA testing (breast), active surveillance (prostate),\n    smoking cessation (lung), CEA monitoring (colorectal)\n• Comprehensive unit tests for all cancer type staging",
     "AI/Backend Team", "Not Started", "P1 — High"],

    ["Phase 4: AI Integration — Auto-Population", "Sprint 12", "Week 23–24",
     "Data Population Service (AI → Patient Records)",
     "• Port DataPopulationService from Django → NestJS\n• Auto-populate patient records from AI-extracted data:\n  — Vitals records (data_source='automated', source_report FK)\n  — Medication records (data_source='automated')\n  — Allergy records (data_source='automated')\n  — ChronicCondition records (data_source='automated')\n• Report approval workflow:\n  — Admin reviews AI analysis → clicks Approve\n  — Triggers DataPopulationService.populatePatientRecords()\n  — Sets report status to 'approved'\n• Data source tracking on all auto-created records\n• Duplicate detection (don't create duplicate allergies/meds)",
     "Backend Team", "Not Started", "P1 — High"],

    ["Phase 4: AI Integration — Auto-Population", "Sprint 12", "Week 23–24",
     "Auto Oncology Record Creation from Pathology Reports",
     "• Port auto_create_oncology_record from Django → NestJS\n• Trigger: pathology report with recognized cancer_type in extracted_data\n• Auto-create OncologyRecord with:\n  — Mapped cancer type (breast → breast_cancer, etc.)\n  — TNM staging from AI extraction\n  — Histology, grade, tumor size, biomarkers\n  — confirmed_by = 'ai', diagnosis_confirmed = false\n  — AI confidence score\n• Link PathologyReport FK to OncologyRecord\n• POST /api/oncology-records/create-from-report endpoint",
     "Backend Team", "Not Started", "P1 — High"],

    ["Phase 4: AI Integration — Treatment Roadmap", "Sprint 12", "Week 23–24",
     "AI Treatment Roadmap Generation",
     "• Port treatment roadmap generation from Django → NestJS\n• Second Groq API call after oncology record creation\n• Prompt includes: cancer type, AJCC stage, histology, grade,\n  tumor size, lymph node status, biomarkers, AI analysis insights\n• LLM generates NCCN-guideline-based recommendations:\n  — recommended_surgery (specific procedure suggestions)\n  — recommended_chemotherapy (specific regimen suggestions)\n  — recommended_radiation (type + dosing)\n  — recommended_immunotherapy (specific agents)\n  — recommended_targeted_therapy (mutation-specific agents)\n  — treatment_intent (curative/palliative/adjuvant/neoadjuvant)\n  — urgency_level (emergent/urgent/semi_urgent/routine)\n• Save roadmap to OncologyRecord fields\n• roadmap_generated_at timestamp\n• GET /api/oncology-records/:id/treatment-roadmap endpoint\n• Integration tests with mock Groq responses",
     "AI/Backend Team", "Not Started", "P1 — High"],

    # ──────────────────────────────────────────────────────────────
    # PHASE 5: AI PROCESSING PIPELINE & PROGRESS TRACKING
    # ──────────────────────────────────────────────────────────────
    ["Phase 5: AI Processing Pipeline", "Sprint 13", "Week 25–26",
     "Document Processing Pipeline (End-to-End)",
     "• Port DocumentProcessingService orchestrator to NestJS\n• Full processing pipeline:\n  Step 1: File upload → validate → store\n  Step 2: TextExtractionService.extractText()\n    → PDF? → pdfplumber/pdf-parse extraction\n    → Scanned PDF? → pdf2image + Tesseract OCR\n    → Image? → preprocessing + Tesseract OCR\n    → Save extracted_text, status='text_extracted'\n  Step 3: AIAnalysisService.analyzeMedicalText()\n    → Token estimation → direct or chunked mode\n    → Post-processing: abnormal lab detection, AJCC staging\n    → Save extracted_data, insights, recommendations\n    → status='analysis_complete'\n  Step 3b (if pathology): Auto-create OncologyRecord\n    → Generate treatment roadmap (2nd Groq API call)\n  Step 4: Await admin approval\n  Step 5: DataPopulationService.populatePatientRecords()\n• Background processing with BullMQ (Redis-backed queue)\n• Configurable concurrency & retry policies",
     "AI/Backend Team", "Not Started", "P1 — High"],

    ["Phase 5: AI Processing Pipeline", "Sprint 13", "Week 25–26",
     "Processing Progress Tracking (Real-Time)",
     "• ProcessingProgress entity tracking stages:\n  initializing → extracting_text → running_ocr → chunking →\n  ai_analysis → merging_results → staging → populating → complete/failed\n• Current step / total steps / percentage tracking\n• Current page / current chunk granular progress\n• WebSocket or SSE for real-time progress updates to frontend\n• GET /api/medical-reports/:id/progress endpoint\n• Progress bar component in React frontend\n• Error recovery: resume from last successful step",
     "Full Stack Team", "Not Started", "P2 — Medium"],

    ["Phase 5: AI Processing Pipeline", "Sprint 14", "Week 27–28",
     "AI Quality & Confidence Scoring",
     "• Document complexity analyzer (port from Django)\n  — Estimate processing difficulty based on page count, OCR quality,\n    document type, text density\n• AI confidence scoring per extraction field\n• Low-confidence flagging for human review\n• A/B comparison: AI-extracted vs manually-entered data\n• AI extraction accuracy dashboard for admins\n• Bulk document processing (batch upload)\n• Processing queue management (pause, cancel, re-priority)",
     "AI/Backend Team", "Not Started", "P2 — Medium"],

    # ──────────────────────────────────────────────────────────────
    # PHASE 6: CHEMO & TREATMENT ENGINE
    # ──────────────────────────────────────────────────────────────
    ["Phase 6: Chemo & Treatment Engine", "Sprint 15", "Week 29–30",
     "Chemotherapy Scheduling Engine",
     "• Calendar-based chemo cycle scheduling\n• Auto-calculate next session dates based on protocol\n• Missed session detection & flagging\n• Pre-chemo checklist (blood work, vitals)\n• Cycle completion tracking (X of Y done)\n• Treatment delay documentation\n• AI-suggested optimal scheduling based on patient data",
     "Backend Lead", "Not Started", "P1 — High"],

    ["Phase 6: Chemo & Treatment Engine", "Sprint 15", "Week 29–30",
     "Medication & Prescription Management",
     "• Chemo drug catalog (per cancer type)\n• Dosage calculation (BSA-based)\n• Supportive medication tracking (anti-nausea, etc.)\n• Extra/emergency medication logging\n• Drug interaction alerts\n• Chemist/Pharmacist verification workflow\n• AI-extracted medications cross-reference with prescriptions",
     "Backend + Chemist Flow", "Not Started", "P1 — High"],

    # ──────────────────────────────────────────────────────────────
    # PHASE 7: NOTIFICATIONS & ALERTS
    # ──────────────────────────────────────────────────────────────
    ["Phase 7: Notifications & Alerts", "Sprint 16", "Week 31–32",
     "Notification System",
     "• Email notifications (SendGrid/AWS SES)\n• SMS notifications (Twilio)\n• Push notifications (Firebase)\n• Notification templates per event type\n• Notification preferences per user\n• Notification history & read tracking\n• AI processing completion notifications",
     "Backend Team", "Not Started", "P1 — High"],

    ["Phase 7: Notifications & Alerts", "Sprint 16", "Week 31–32",
     "Alert Rules Engine",
     "• Missed chemo session → Notify patient + emergency contact\n• Doctor unavailable → Notify patient + assign backup\n• Abnormal benchmark → Alert care team\n• ER visit logged → Notify oncologist\n• Treatment milestone → Notify patient\n• Custom alert rules per hospital\n• AI-detected critical findings → Urgent alert to oncologist",
     "Backend Team", "Not Started", "P1 — High"],

    # ──────────────────────────────────────────────────────────────
    # PHASE 8: BENCHMARKS, ANALYTICS & PATIENT PORTAL
    # ──────────────────────────────────────────────────────────────
    ["Phase 8: Benchmarks & Analytics", "Sprint 17", "Week 33–34",
     "Benchmarking Dashboard",
     "• Tumor marker trend charts (up/down arrows)\n• Treatment response rates per cancer type\n• Chemo completion rates\n• Side-effect frequency tracking\n• Comparative benchmarks (hospital vs national)\n• Patient outcome scoring\n• AI-extracted lab result trends over time",
     "Frontend + Data", "Not Started", "P1 — High"],

    ["Phase 8: Benchmarks & Analytics", "Sprint 17", "Week 33–34",
     "ER & Emergency Tracking",
     "• Emergency room visit logging\n• Reason categorization (fever, pain, reaction, etc.)\n• Extra medications administered in ER\n• Unplanned hospitalization tracking\n• ER visit trends per patient/cancer type\n• Cost impact analysis",
     "Full Stack Team", "Not Started", "P1 — High"],

    ["Phase 8: Benchmarks & Analytics", "Sprint 18", "Week 35–36",
     "Patient-Facing Portal",
     "• Treatment journey timeline view\n• Remaining chemo sessions counter\n• Next appointment details\n• Medication list & schedule\n• Doctor/nurse contact info\n• Symptom self-reporting\n• AI-generated treatment summary (plain language)\n• Document access (reports, prescriptions)\n• Appointment reminders (email + SMS + push)",
     "Frontend Team", "Not Started", "P1 — High"],

    # ──────────────────────────────────────────────────────────────
    # PHASE 9: REPORTING & COMPLIANCE
    # ──────────────────────────────────────────────────────────────
    ["Phase 9: Reporting & Compliance", "Sprint 19", "Week 37–38",
     "Reports & Exports",
     "• Patient treatment summary reports (PDF)\n• Hospital-wide oncology analytics\n• Insurance/billing report generation\n• Compliance audit reports\n• AI analysis audit trail (who approved, what was changed)\n• Custom report builder\n• Scheduled report delivery (email)",
     "Full Stack Team", "Not Started", "P2 — Medium"],

    # ──────────────────────────────────────────────────────────────
    # PHASE 10: TESTING & DEPLOYMENT
    # ──────────────────────────────────────────────────────────────
    ["Phase 10: Testing & Deployment", "Sprint 20–21", "Week 39–42",
     "QA, UAT & Deployment",
     "• Unit & integration testing (>80% coverage)\n• End-to-end testing (Cypress/Playwright)\n• AI service testing (mock Groq responses, OCR test images)\n• Security audit & penetration testing\n• HIPAA compliance review\n• Performance & load testing\n• AI processing load testing (concurrent document uploads)\n• UAT with hospital staff\n• Production deployment\n• Documentation & training materials",
     "QA + DevOps", "Not Started", "P0 — Critical"],
]

for r, row in enumerate(phases, 1):
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = DARK_FILL
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL
            # Color-code the Phase column
            if c == 1 and val:
                if "Infrastructure" in str(val):
                    cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, size=11, color="065F46")
                elif "AI" in str(val):
                    cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, size=11, color="3730A3")
            # Color-code priority
            if c == 8 and val:
                if "P0" in str(val):
                    cell.fill = RED_FILL
                    cell.font = Font(name="Calibri", bold=True, color="991B1B")
                elif "P1" in str(val):
                    cell.fill = YELLOW_FILL
                    cell.font = Font(name="Calibri", bold=True, color="92400E")
                elif "P2" in str(val):
                    cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")

ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 40
ws2.column_dimensions["E"].width = 65
ws2.column_dimensions["F"].width = 22
ws2.column_dimensions["G"].width = 14
ws2.column_dimensions["H"].width = 16
ws2.row_dimensions[1].height = 30
for r in range(2, len(phases) + 1):
    ws2.row_dimensions[r].height = 130


# ═══════════════════════════════════════════════════════════════════
# TAB 3 — AI FEATURES DETAIL
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. AI Features Detail")
ws3.sheet_properties.tabColor = "6366F1"

ai_data = [
    ["AI Feature", "What It Does", "Input", "Output", "Technology", "Django Reference", "NestJS Implementation Plan"],

    ["DOCUMENT PROCESSING PIPELINE", "", "", "", "", "", ""],

    ["PDF Text Extraction",
     "Extracts text from born-digital PDFs (text is selectable in the PDF). Uses pdfplumber to directly read text layer.",
     "PDF file (born-digital)",
     "Plain text (extracted_text field on MedicalReport)",
     "pdfplumber (Python)\n→ pdf-parse or pdf.js (Node)",
     "text_extraction_service.py\nLines 260-285",
     "Create TextExtractionService in NestJS\nUse pdf-parse npm package\nSame logic: extract text per page, concatenate"],

    ["Scanned PDF Detection",
     "Detects if a PDF is image-based (scanned) by checking if first page has < 50 characters of extractable text.",
     "PDF file",
     "Boolean (is_scanned)",
     "pdfplumber text length check",
     "text_extraction_service.py\nLines 55-70",
     "After pdf-parse extraction, check text.length < 50\nIf true → route to OCR pipeline"],

    ["Scanned PDF OCR",
     "Converts each PDF page to an image, then runs Tesseract OCR on each page. Handles multi-page scanned documents.",
     "Scanned PDF file",
     "Extracted text (per page, concatenated)",
     "pdf2image + pytesseract + PIL\n→ pdf-poppler/pdf2pic + node-tesseract-ocr + sharp (Node)",
     "text_extraction_service.py\nLines 105-175",
     "Use pdf-poppler or pdf2pic to convert PDF → images\nUse node-tesseract-ocr or tesseract.js for OCR\nProcess page by page"],

    ["Image OCR with Preprocessing",
     "Advanced image preprocessing before OCR: convert to grayscale, enhance contrast (2.0x), sharpen, upscale small images. Then run Tesseract with optimized config.",
     "Image file (PNG/JPG/JPEG/TIFF)",
     "Extracted text (cleaned)",
     "PIL (contrast/sharpen/resize) + pytesseract\n→ sharp (Node) + tesseract.js",
     "text_extraction_service.py\nLines 290-345",
     "Use sharp for image preprocessing:\n- .grayscale()\n- .sharpen()\n- .resize() for small images\nTesseract config: --oem 3 --psm 3"],

    ["OCR Text Cleanup",
     "Post-processes OCR output to fix common artifacts: whitespace normalization, pipe→I replacement, medical term corrections (Gleason, Adenocarcinoma), noise line removal.",
     "Raw OCR text",
     "Cleaned text",
     "Regex (Python re)\n→ Regex (Node.js)",
     "text_extraction_service.py\nLines 347-374",
     "Port regex patterns to Node.js\nSame cleanup rules:\n- Fix |→I\n- Fix spacing\n- Medical term corrections"],

    ["", "", "", "", "", "", ""],
    ["AI ANALYSIS (LLM-POWERED)", "", "", "", "", "", ""],

    ["Medical Report AI Analysis (Direct)",
     "Sends full extracted text to Groq Llama 3.3 70B in a single API call. LLM extracts structured medical data and generates clinical insights + recommendations.",
     "Extracted text (≤ 6000 tokens)\n+ Document type",
     "JSON: {\n  vitals, lab_results,\n  medications, allergies,\n  diagnoses, insights,\n  recommendations,\n  oncology_data (if pathology)\n}",
     "Groq API (llama-3.3-70b-versatile)\ntemp=0.1, max_tokens=1500\nJSON response format",
     "ai_analysis_service.py\nLines 52-167",
     "Use groq-sdk npm package\nSame API params: model, temp, max_tokens\nSame prompt structure per document type\nParse JSON response with validation"],

    ["Medical Report AI Analysis (Chunked)",
     "For large documents (> 6000 tokens): splits text into overlapping chunks, sends each to Groq separately, then intelligently merges all results.",
     "Extracted text (> 6000 tokens)\n+ Document type",
     "Merged JSON (same structure as direct)",
     "ChunkingService + Groq API\nN API calls (one per chunk)",
     "ai_analysis_service.py\nLines 169-262",
     "Port chunking logic to NestJS\n6000 token chunks, 200 overlap\nMerge: first-non-null for scalars,\nunion for lists, merge for dicts"],

    ["Document Type Auto-Detection",
     "Rule-based NLP that classifies document type from text content using keyword matching. NO AI API call — pure regex. Detects: pathology, ultrasound, X-ray, CT, MRI, PET, mammography, radiology, lab report.",
     "Extracted text (any document)",
     "Document type classification string",
     "Keyword matching + counting\nNo LLM / No API call",
     "ai_analysis_service.py\nLines 310-370",
     "Port keyword dictionaries to NestJS\nSame logic: count keyword matches\n≥2 matches for pathology (OCR robustness)"],

    ["Pathology Report Deep Extraction",
     "Specialized LLM prompt for pathology reports that extracts the FULL oncology schema: cancer type, TNM staging, histology, biomarkers (ER/PR/HER2, Gleason, EGFR, ALK, PD-L1, KRAS, BRAF, MSI/MMR), specimen info, ICD codes, immunohistochemistry.",
     "Pathology report text + pathology-specific prompt",
     "JSON: {\n  oncology_data: {\n    cancer_type, tnm_staging,\n    histology, grade,\n    tumor_size, biomarkers,\n    lymph_nodes, margins,\n    specimen_info, icd_codes,\n    ihc_results\n  },\n  pathology_info: { ... }\n}",
     "Groq API with specialized\npathology prompt (215 lines)",
     "ai_analysis_service.py\nLines 855-1070",
     "Port the 215-line pathology prompt\nJSON schema must match exactly\nThis is the most complex prompt —\ntest with real pathology reports"],

    ["Abnormal Lab Detection",
     "Post-processes AI-extracted lab results. Parses reference ranges and flags values outside normal (is_abnormal). Handles edge cases: '<', '>', ranges '10-20'. Deterministic — NOT an AI call.",
     "AI-extracted lab results array",
     "Lab results with is_abnormal boolean",
     "Deterministic parsing\nNo LLM / No API call",
     "ai_analysis_service.py\nLines 264-291",
     "Port parsing logic to TypeScript\nHandle: numeric, ranges, '<', '>'\nSet is_abnormal on each result"],

    ["", "", "", "", "", "", ""],
    ["STAGING & RECOMMENDATIONS", "", "", "", "", "", ""],

    ["AJCC 8th Edition Staging (Breast)",
     "Deterministic staging engine for breast cancer. Anatomic staging (T+N+M) and Prognostic staging (T+N+M+Grade+ER/PR/HER2). Classifies molecular subtypes: Luminal A/B, HER2+, Triple-negative.",
     "TNM values, grade,\nER/PR/HER2 status",
     "AJCC stage (IA, IB, IIA, IIB, IIIA, IIIB, IIIC, IV)\n+ Molecular subtype\n+ Prognostic stage",
     "Lookup tables\nNo LLM / No API call",
     "ajcc_staging_engine.py\nLines 85-260",
     "Port staging lookup tables to TypeScript\nMap interface for each staging input\nReturn structured staging result"],

    ["AJCC 8th Edition Staging (Prostate)",
     "NCCN Risk Group classification for prostate cancer based on T stage, PSA level, and Gleason score (primary + secondary).",
     "T stage, PSA (ng/mL),\nGleason primary + secondary",
     "Risk group: Very Low, Low,\nFavorable/Unfavorable Intermediate,\nHigh, Very High, Metastatic",
     "Decision tree / lookup\nNo LLM / No API call",
     "ajcc_staging_engine.py\nLines 265-385",
     "Port decision tree to TypeScript\nPSA thresholds: <10, 10-20, >20\nGleason groups: 3+3, 3+4, 4+3, ≥8"],

    ["AJCC 8th Edition Staging (Lung NSCLC)",
     "TNM-based stage grouping for non-small cell lung cancer.",
     "T, N, M stage values",
     "Stage: IA1-IA3, IB, IIA, IIB,\nIIIA, IIIB, IIIC, IVA, IVB",
     "Lookup table\nNo LLM / No API call",
     "ajcc_staging_engine.py\nLines 390-470",
     "Port TNM→Stage lookup table\nHandle all T1-T4, N0-N3, M0-M1 combos"],

    ["AJCC 8th Edition Staging (Colorectal)",
     "TNM-based stage grouping for colorectal cancer.",
     "T, N, M stage values",
     "Stage: 0, I, IIA, IIB, IIC,\nIIIA, IIIB, IIIC, IVA, IVB, IVC",
     "Lookup table\nNo LLM / No API call",
     "ajcc_staging_engine.py\nLines 475-550",
     "Port TNM→Stage lookup table\nHandle Tis, T1-T4, N0-N2, M0-M1"],

    ["AI Treatment Roadmap Generation",
     "After oncology record creation, makes a SECOND Groq API call to generate comprehensive treatment recommendations based on NCCN guidelines. Fills in surgery, chemo, radiation, immunotherapy, targeted therapy fields.",
     "Cancer type, AJCC stage,\nhistology, grade, tumor size,\nlymph nodes, biomarkers,\nAI analysis insights",
     "JSON: {\n  recommended_surgery,\n  recommended_chemotherapy,\n  recommended_radiation,\n  recommended_immunotherapy,\n  recommended_targeted_therapy,\n  treatment_intent,\n  urgency_level\n}",
     "Groq API (2nd call)\nNNCN-guideline-aware prompt",
     "data_population_service.py\nLines 370-495",
     "Port roadmap generation prompt\nSave to OncologyRecord fields\nSet roadmap_generated_at timestamp\nDedicated endpoint to view roadmap"],

    ["Stage-Specific Recommendations",
     "Generates stage-specific and cancer-type-specific recommendations after AJCC staging. E.g., Stage IV → urgent systemic therapy; breast → BRCA testing.",
     "AJCC stage + cancer type",
     "Array of recommendation strings\nadded to insights/recommendations",
     "Rule-based / deterministic\nNo LLM call",
     "ai_analysis_service.py\nLines 1180-1239",
     "Port recommendation rules to TypeScript\nStage IV/III/II rules\nCancer-specific rules (BRCA, PSA, etc.)"],

    ["", "", "", "", "", "", ""],
    ["AUTO-POPULATION FLOWS", "", "", "", "", "", ""],

    ["Auto Oncology Record from Pathology",
     "When a pathology report finishes AI analysis, automatically creates an OncologyRecord with extracted data. Sets confirmed_by='ai', diagnosis_confirmed=false (requires doctor approval).",
     "AI-analyzed pathology report\nwith oncology_data containing\nrecognized cancer_type",
     "New OncologyRecord linked\nto pathology report\n+ Treatment roadmap (via AI)",
     "Business logic + Groq API\n(for roadmap generation)",
     "data_population_service.py\nLines 218-370",
     "Port auto-creation logic\nMap cancer types\nRequire doctor confirmation\nTrigger roadmap generation"],

    ["Auto Patient Record Population",
     "On report approval, auto-populates patient's Vitals, Medications, Allergies, ChronicConditions from AI-extracted data. All records tagged with data_source='automated' and linked to source report.",
     "Approved MedicalReport\nwith extracted_data",
     "New Vitals, Medication,\nAllergy, ChronicCondition\nrecords (data_source='automated')",
     "Business logic\nNo AI call — uses previously\nextracted data",
     "data_population_service.py\nLines 30-215",
     "Port population logic\nDuplicate detection\nSet data_source='automated'\nSet source_report FK"],

    ["", "", "", "", "", "", ""],
    ["PROCESSING INFRASTRUCTURE", "", "", "", "", "", ""],

    ["Processing Progress Tracker",
     "Tracks document processing through stages with real-time progress: initializing → extracting_text → running_ocr → chunking → ai_analysis → merging_results → staging → complete/failed",
     "Processing events from\neach pipeline stage",
     "ProcessingProgress entity:\ncurrent_step, total_steps,\npercentage, current_page,\ncurrent_chunk, status",
     "Database entity + WebSocket/SSE",
     "ProcessingProgress model\nProgress tracker service",
     "ProcessingProgress entity\nProgress service with event emitter\nSSE endpoint or WebSocket gateway\nReact progress bar component"],

    ["Document Complexity Analyzer",
     "Estimates processing difficulty/time based on page count, file size, OCR quality, document type, text density. Used for queue prioritization and user time estimates.",
     "Document metadata:\npage_count, file_size,\nis_scanned, document_type",
     "Complexity score\nEstimated processing time",
     "Heuristic scoring\nNo AI call",
     "document_complexity_analyzer.py",
     "Port scoring heuristics\nUse for BullMQ priority\nShow estimated time in UI"],
]

for r, row in enumerate(ai_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = AI_FILL
        elif val in ("DOCUMENT PROCESSING PIPELINE", "AI ANALYSIS (LLM-POWERED)",
                     "STAGING & RECOMMENDATIONS", "AUTO-POPULATION FLOWS",
                     "PROCESSING INFRASTRUCTURE"):
            cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
            for cc in range(2, 8):
                ws3.cell(row=r, column=cc).fill = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL

ws3.column_dimensions["A"].width = 35
ws3.column_dimensions["B"].width = 50
ws3.column_dimensions["C"].width = 30
ws3.column_dimensions["D"].width = 32
ws3.column_dimensions["E"].width = 30
ws3.column_dimensions["F"].width = 25
ws3.column_dimensions["G"].width = 35
for r in range(2, len(ai_data) + 1):
    ws3.row_dimensions[r].height = 110


# ═══════════════════════════════════════════════════════════════════
# TAB 4 — AI TECH STACK & DEPENDENCIES
# ═══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. AI Tech Stack")
ws4.sheet_properties.tabColor = "4F46E5"

ai_tech_data = [
    ["Layer", "Django (Current)", "NestJS (Target)", "NPM Package", "Purpose", "Notes"],

    ["LLM API", "", "", "", "", ""],
    ["LLM Provider", "Groq Cloud API", "Groq Cloud API (same)", "groq-sdk", "LLM inference for medical text analysis", "Same API key works for both"],
    ["LLM Model", "llama-3.3-70b-versatile", "llama-3.3-70b-versatile (same)", "—", "70B parameter model for medical extraction", "Best quality/speed ratio on Groq"],
    ["LLM Config", "temp=0.1, max_tokens=1500", "temp=0.1, max_tokens=1500 (same)", "—", "Low temperature for deterministic extraction", "JSON response_format enforced"],
    ["Token Chunking", "6000 token chunks, 200 overlap", "6000 token chunks, 200 overlap (same)", "—", "Handle documents exceeding context window", "Character/4 approximation for token count"],

    ["", "", "", "", "", ""],
    ["OCR & PDF", "", "", "", "", ""],
    ["PDF Text Extraction", "pdfplumber 0.10.3", "pdf-parse / pdfjs-dist", "pdf-parse", "Extract text from born-digital PDFs", "pdf-parse is simpler; pdfjs-dist more flexible"],
    ["PDF to Image", "pdf2image 1.16.3 + Poppler", "pdf-poppler / pdf2pic", "pdf-poppler or pdf2pic", "Convert PDF pages to images for OCR", "Requires Poppler system binary installed"],
    ["OCR Engine", "pytesseract 0.3.10 + Tesseract", "node-tesseract-ocr / tesseract.js", "node-tesseract-ocr", "Extract text from images/scanned docs", "node-tesseract-ocr = native Tesseract (faster)\ntesseract.js = WASM (no system dep, slower)"],
    ["Image Processing", "Pillow 10.1.0", "sharp", "sharp", "Image preprocessing: grayscale, contrast, sharpen, resize", "sharp is the Node standard for image processing"],

    ["", "", "", "", "", ""],
    ["SYSTEM DEPENDENCIES (must install on server)", "", "", "", "", ""],
    ["Tesseract OCR", "tesseract-ocr (apt install)", "tesseract-ocr (same)", "—", "OCR engine binary", "apt install tesseract-ocr tesseract-ocr-eng"],
    ["Poppler Utils", "poppler-utils (apt install)", "poppler-utils (same)", "—", "PDF rendering tools (pdftoppm)", "apt install poppler-utils"],
    ["libvips", "N/A", "Required by sharp", "—", "Image processing library", "Usually auto-installed with sharp npm package"],

    ["", "", "", "", "", ""],
    ["QUEUE & PROCESSING", "", "", "", "", ""],
    ["Job Queue", "Synchronous (in-request)", "BullMQ (Redis-backed)", "@nestjs/bullmq, bullmq", "Async background processing for document pipeline", "Allows concurrent processing, retries, priority"],
    ["Redis", "Not used in Django AI", "Required for BullMQ", "ioredis", "Queue backend + caching", "Already in tech stack for sessions"],
    ["Progress Tracking", "ProcessingProgress model", "ProcessingProgress entity + SSE/WS", "@nestjs/event-emitter", "Real-time processing progress updates", "Server-Sent Events or WebSocket Gateway"],

    ["", "", "", "", "", ""],
    ["DETERMINISTIC ENGINES (no external deps)", "", "", "", "", ""],
    ["AJCC Staging", "ajcc_staging_engine.py (581 lines)", "ajcc-staging.service.ts", "—", "Cancer staging lookup tables", "Pure TypeScript — no dependencies"],
    ["Lab Abnormal Detection", "ai_analysis_service.py L264-291", "lab-detection.service.ts", "—", "Reference range parsing + flagging", "Pure TypeScript — no dependencies"],
    ["Report Type Detection", "ai_analysis_service.py L310-370", "report-classifier.service.ts", "—", "Keyword-based document classification", "Pure TypeScript — no dependencies"],
    ["Stage Recommendations", "ai_analysis_service.py L1180-1239", "staging-recommendations.service.ts", "—", "Stage-specific treatment recommendations", "Pure TypeScript — no dependencies"],

    ["", "", "", "", "", ""],
    ["DOCKER IMAGE ADDITIONS", "", "", "", "", ""],
    ["Dockerfile additions", "N/A", "Add to NestJS Dockerfile:", "—", "", ""],
    ["", "", "RUN apt-get update && apt-get install -y \\", "—", "Install system deps in Docker", ""],
    ["", "", "  tesseract-ocr tesseract-ocr-eng \\", "—", "Tesseract + English language pack", ""],
    ["", "", "  poppler-utils \\", "—", "PDF tools (pdftoppm, pdfinfo)", ""],
    ["", "", "  && rm -rf /var/lib/apt/lists/*", "—", "Clean apt cache", ""],
]

for r, row in enumerate(ai_tech_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        elif val in ("LLM API", "OCR & PDF", "SYSTEM DEPENDENCIES (must install on server)",
                     "QUEUE & PROCESSING", "DETERMINISTIC ENGINES (no external deps)",
                     "DOCKER IMAGE ADDITIONS"):
            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
            for cc in range(2, 7):
                ws4.cell(row=r, column=cc).fill = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL

ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 32
ws4.column_dimensions["C"].width = 38
ws4.column_dimensions["D"].width = 28
ws4.column_dimensions["E"].width = 42
ws4.column_dimensions["F"].width = 38


# ═══════════════════════════════════════════════════════════════════
# TAB 5 — ROLES & PERMISSIONS MATRIX
# ═══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5. Roles & Permissions")
ws5.sheet_properties.tabColor = "7C3AED"

roles_header = ["Module / Feature", "SuperAdmin", "Hospital Admin", "Doctor", "Nurse", "Chemist / Pharmacist", "Patient"]

role_data = [
    roles_header,
    ["— USER MANAGEMENT —", "", "", "", "", "", ""],
    ["Create/Edit Hospital", "Full", "—", "—", "—", "—", "—"],
    ["Create/Edit Users (Doctors, Nurses)", "Full", "Full", "—", "—", "—", "—"],
    ["Assign Roles & Permissions", "Full", "Limited", "—", "—", "—", "—"],
    ["View All Users Across Hospitals", "Full", "—", "—", "—", "—", "—"],
    ["Manage Own Profile", "Full", "Full", "Full", "Full", "Full", "Full"],
    ["", "", "", "", "", "", ""],
    ["— PATIENT MANAGEMENT —", "", "", "", "", "", ""],
    ["Register New Patient", "Full", "Full", "Full", "Full", "—", "—"],
    ["View Patient List", "Full", "Full", "Full", "Full", "Limited", "Own Only"],
    ["Edit Patient Demographics", "Full", "Full", "Full", "Limited", "—", "Own Only"],
    ["View Patient Medical History", "Full", "Full", "Full", "Full", "Limited", "Own Only"],
    ["Manage Emergency Contacts", "Full", "Full", "Full", "Full", "—", "Own Only"],
    ["Upload/View Documents", "Full", "Full", "Full", "Full", "Limited", "Own Only"],
    ["", "", "", "", "", "", ""],
    ["— AI DOCUMENT ANALYSIS —", "", "", "", "", "", ""],
    ["Upload Medical Reports for AI Analysis", "Full", "Full", "Full", "Full", "—", "—"],
    ["View AI Processing Progress", "Full", "Full", "Full", "Full", "—", "—"],
    ["View AI Extracted Data & Insights", "Full", "Full", "Full", "Full", "Limited", "Own Only"],
    ["Approve AI Analysis (populate records)", "Full", "Full", "Full", "—", "—", "—"],
    ["Retry Failed AI Processing", "Full", "Full", "Full", "—", "—", "—"],
    ["View AI Treatment Roadmap", "Full", "Full", "Full", "Full", "View", "Own Only"],
    ["Confirm AI-Created Oncology Record", "—", "—", "Full", "—", "—", "—"],
    ["View AI Confidence Scores", "Full", "Full", "Full", "Full", "—", "—"],
    ["AI Analysis Audit Trail", "Full", "Full", "—", "—", "—", "—"],
    ["", "", "", "", "", "", ""],
    ["— ONCOLOGY / CANCER MODULES —", "", "", "", "", "", ""],
    ["Create Cancer Diagnosis Record", "—", "—", "Full", "—", "—", "—"],
    ["Edit Diagnosis Details", "Full", "—", "Full", "—", "—", "—"],
    ["View Diagnosis Details", "Full", "Full", "Full", "Full", "View", "Own Only"],
    ["Assign Chemo Protocol", "—", "—", "Full", "—", "—", "—"],
    ["View Treatment Plan", "Full", "Full", "Full", "Full", "Full", "Own Only"],
    ["View AJCC Staging Results", "Full", "Full", "Full", "Full", "View", "Own Only"],
    ["", "", "", "", "", "", ""],
    ["— CHEMOTHERAPY TRACKING —", "", "", "", "", "", ""],
    ["Schedule Chemo Sessions", "—", "—", "Full", "Full", "—", "—"],
    ["Mark Session Complete", "—", "—", "Full", "Full", "—", "—"],
    ["Log Missed Session", "—", "—", "Full", "Full", "—", "—"],
    ["View Chemo Calendar", "Full", "Full", "Full", "Full", "Full", "Own Only"],
    ["View Remaining Sessions Count", "Full", "Full", "Full", "Full", "Full", "Own Only"],
    ["Pre-Chemo Checklist", "—", "—", "Full", "Full", "—", "—"],
    ["", "", "", "", "", "", ""],
    ["— MEDICATIONS —", "", "", "", "", "", ""],
    ["Prescribe Medications", "—", "—", "Full", "—", "—", "—"],
    ["Verify/Dispense Medications", "—", "—", "—", "—", "Full", "—"],
    ["View Medication List", "Full", "Full", "Full", "Full", "Full", "Own Only"],
    ["Log Extra/Emergency Meds", "—", "—", "Full", "Full", "Full", "—"],
    ["Drug Interaction Alerts", "—", "—", "View", "View", "Full", "—"],
    ["", "", "", "", "", "", ""],
    ["— NOTIFICATIONS —", "", "", "", "", "", ""],
    ["Configure Notification Rules", "Full", "Full", "—", "—", "—", "—"],
    ["Send Manual Notifications", "Full", "Full", "Full", "Full", "—", "—"],
    ["Receive Treatment Alerts", "—", "—", "Full", "Full", "Full", "Full"],
    ["Manage Notification Preferences", "Full", "Full", "Full", "Full", "Full", "Full"],
    ["View Notification History", "Full", "Full", "Own", "Own", "Own", "Own"],
    ["", "", "", "", "", "", ""],
    ["— BENCHMARKS & ANALYTICS —", "", "", "", "", "", ""],
    ["View Hospital-Wide Analytics", "Full", "Full", "—", "—", "—", "—"],
    ["View Patient Benchmarks", "Full", "Full", "Full", "Full", "—", "Own Only"],
    ["View Tumor Marker Trends", "Full", "Full", "Full", "Full", "—", "Own Only"],
    ["Compare Hospital vs National", "Full", "Full", "Full", "—", "—", "—"],
    ["Export Reports", "Full", "Full", "Full", "Limited", "—", "—"],
    ["", "", "", "", "", "", ""],
    ["— ER & EMERGENCY —", "", "", "", "", "", ""],
    ["Log ER Visit", "—", "—", "Full", "Full", "—", "—"],
    ["View ER History", "Full", "Full", "Full", "Full", "—", "Own Only"],
    ["Log Emergency Medications", "—", "—", "Full", "Full", "Full", "—"],
    ["ER Trend Analytics", "Full", "Full", "Full", "—", "—", "—"],
    ["", "", "", "", "", "", ""],
    ["— REPORTS —", "", "", "", "", "", ""],
    ["Generate Patient Reports", "Full", "Full", "Full", "Limited", "—", "Own Only"],
    ["Generate Hospital Reports", "Full", "Full", "—", "—", "—", "—"],
    ["Insurance/Billing Reports", "Full", "Full", "Full", "—", "—", "Own Only"],
    ["Audit Trail Reports", "Full", "Full", "—", "—", "—", "—"],
]

for r, row in enumerate(role_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws5.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = PURPLE_FILL
        elif val and val.startswith("—") and val.endswith("—"):
            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        elif val == "Full":
            cell.fill = GREEN_FILL
            cell.font = Font(name="Calibri", bold=True, size=11, color="065F46")
        elif val == "Limited" or val == "View":
            cell.fill = YELLOW_FILL
            cell.font = Font(name="Calibri", size=11, color="92400E")
        elif val == "Own Only" or val == "Own":
            cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
            cell.font = Font(name="Calibri", size=11, color="3730A3")
        elif val == "—":
            cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            cell.font = Font(name="Calibri", size=11, color="9CA3AF")
        else:
            cell.font = BODY_FONT

ws5.column_dimensions["A"].width = 42
for col_letter in ["B", "C", "D", "E", "F", "G"]:
    ws5.column_dimensions[col_letter].width = 22


# ═══════════════════════════════════════════════════════════════════
# TAB 6 — CANCER MODULES DETAIL
# ═══════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6. Cancer Modules Detail")
ws6.sheet_properties.tabColor = "DC2626"

cancer_data = [
    ["Cancer Type", "Diagnosis Fields", "Chemo Protocols", "Key Tumor Markers", "Benchmark Metrics", "AI Extraction Focus", "AJCC Staging Logic"],
    [
        "Breast Cancer",
        "• Tumor Size & Grade\n• ER/PR Status (+/-)\n• HER2 Status (+/-)\n• BRCA1/BRCA2 mutation\n• Lymph node involvement\n• TNM Staging (I-IV)\n• Histology type\n• Ki-67 index",
        "• AC-T (Adriamycin + Cyclophosphamide → Taxol)\n• TC (Taxotere + Cyclophosphamide)\n• TCH (for HER2+)\n• CMF (Cyclophosphamide + Methotrexate + 5-FU)\n• Trastuzumab (Herceptin)\n• Hormone Therapy (Tamoxifen/AI)",
        "• CA 15-3\n• CA 27.29\n• CEA\n• HER2/neu levels",
        "• CA 15-3 trend (↑↓)\n• Tumor size reduction %\n• Chemo completion rate\n• Recurrence-free survival\n• Treatment response (CR/PR/SD/PD)",
        "AI extracts from pathology:\n• ER/PR/HER2 (IHC scores)\n• Ki-67 percentage\n• Tumor size in mm/cm\n• Lymph node count (+/-)\n• Grade (Nottingham)\n• Histology type\n• TNM staging\n• BRCA status if noted\n• Molecular subtype",
        "Anatomic: T+N+M\nPrognostic: T+N+M+Grade+ER/PR/HER2\nSubtypes:\n• Luminal A (ER+/HER2-/low Ki67)\n• Luminal B (ER+/HER2-/high Ki67)\n• HER2+ (HER2+)\n• Triple-neg (ER-/PR-/HER2-)"
    ],
    [
        "Prostate Cancer",
        "• Gleason Score (3+3 to 5+5)\n• PSA Level (ng/mL)\n• TNM Staging\n• Risk Group (Low/Inter/High)\n• Biopsy results\n• Digital rectal exam\n• Bone scan results",
        "• Docetaxel + Prednisone\n• Cabazitaxel\n• Mitoxantrone\n• Abiraterone + Prednisone\n• Enzalutamide\n• ADT (LHRH agonists/antagonists)\n• Radiation combos",
        "• PSA (ng/mL)\n• PSA Velocity\n• PSA Doubling Time\n• Testosterone levels\n• Alkaline Phosphatase",
        "• PSA trend (↑↓) over time\n• PSA response (>50% drop)\n• Time to PSA nadir\n• Chemo completion rate\n• Bone metastasis progression",
        "AI extracts from pathology:\n• Gleason primary + secondary\n• PSA at diagnosis\n• Number of +/- cores\n• Tumor volume %\n• Perineural invasion\n• Extraprostatic extension\n• Seminal vesicle invasion\n• Margin status",
        "NCCN Risk Groups:\n• Very Low: T1c, Gleason 3+3, PSA<10, <3 cores+\n• Low: T1-T2a, Gleason 3+3, PSA<10\n• Intermediate: T2b-T2c or Gleason 3+4/4+3 or PSA 10-20\n• High: T3a or Gleason 8+ or PSA>20\n• Very High: T3b-T4 or primary Gleason 5\n• Metastatic: M1"
    ],
    [
        "Lung Cancer",
        "• Type: SCLC vs NSCLC\n• Subtype (Adeno/Squamous/Large)\n• Staging (I-IV / Limited-Extensive)\n• Mutations (EGFR, ALK, ROS1, PD-L1%)\n• Pulmonary Function Tests\n• Biopsy/Pathology\n• Performance Status (ECOG)",
        "• Cisplatin + Etoposide (SCLC)\n• Carboplatin + Paclitaxel\n• Carboplatin + Pemetrexed (non-squamous)\n• Pembrolizumab (immunotherapy)\n• Atezolizumab + Chemo\n• Targeted: Osimertinib, Crizotinib\n• Concurrent chemoradiation",
        "• CYFRA 21-1\n• NSE (for SCLC)\n• CEA\n• SCC Antigen\n• PD-L1 expression %",
        "• Tumor marker trends (↑↓)\n• PFT trend (FEV1, DLCO)\n• Treatment response (RECIST)\n• Progression-free survival\n• Immunotherapy response rate",
        "AI extracts from pathology:\n• NSCLC vs SCLC classification\n• EGFR mutation status & type\n• ALK rearrangement (IHC/FISH)\n• ROS1 rearrangement\n• PD-L1 TPS percentage\n• KRAS G12C status\n• Histology subtype\n• TNM staging\n• Lymph node stations",
        "NSCLC TNM Staging:\n• IA1-IA3: T1a-c, N0, M0\n• IB: T2a, N0, M0\n• IIA: T2b, N0, M0\n• IIB: T1-2, N1 or T3, N0\n• IIIA: T1-2 N2 or T3 N1 or T4 N0-1\n• IIIB: T1-2 N3 or T3-4 N2\n• IIIC: T3-4, N3\n• IVA-IVB: M1a-c"
    ],
    [
        "Colorectal Cancer",
        "• Tumor Location (Colon/Rectum/Segment)\n• TNM Staging (I-IV)\n• MSI Status (MSI-H/MSS)\n• KRAS/NRAS/BRAF mutation\n• Grade (Well/Mod/Poorly diff)\n• Lymph node count\n• Colonoscopy findings",
        "• FOLFOX (5-FU + Leucovorin + Oxaliplatin)\n• FOLFIRI (5-FU + Leucovorin + Irinotecan)\n• CAPOX (Capecitabine + Oxaliplatin)\n• Bevacizumab (Avastin)\n• Cetuximab (for KRAS wild-type)\n• Pembrolizumab (for MSI-H)\n• Adjuvant vs Palliative regimens",
        "• CEA (ng/mL)\n• CA 19-9\n• LDH\n• Microsatellite status",
        "• CEA trend (↑↓)\n• Treatment response (RECIST)\n• Chemo completion rate\n• Neuropathy grade tracking\n• 5-year survival benchmarks",
        "AI extracts from pathology:\n• Tumor site (cecum→rectum)\n• MSI/MMR status (IHC + PCR)\n• KRAS exon 2/3/4 mutation\n• NRAS mutation status\n• BRAF V600E mutation\n• CEA at diagnosis\n• Lymph nodes examined vs +\n• Tumor deposits\n• Circumferential margin\n• Lymphovascular invasion",
        "TNM Stage Groups:\n• 0: Tis, N0, M0\n• I: T1-T2, N0, M0\n• IIA: T3, N0, M0\n• IIB: T4a, N0, M0\n• IIC: T4b, N0, M0\n• IIIA: T1-2 N1 or T1 N2a\n• IIIB: T3-4a N1 or T2-3 N2a or T1-2 N2b\n• IIIC: T4a N2a or T3-4a N2b or T4b N1-2\n• IVA: M1a\n• IVB: M1b\n• IVC: M1c"
    ],
]

for r, row in enumerate(cancer_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws6.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL

ws6.column_dimensions["A"].width = 20
for col_letter in ["B", "C", "D", "E", "F", "G"]:
    ws6.column_dimensions[col_letter].width = 42
for r in range(2, len(cancer_data) + 1):
    ws6.row_dimensions[r].height = 230


# ═══════════════════════════════════════════════════════════════════
# TAB 7 — NOTIFICATION SYSTEM
# ═══════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("7. Notification System")
ws7.sheet_properties.tabColor = "EA580C"

notif_data = [
    ["#", "Trigger Event", "Channel", "Primary Recipient", "Secondary Recipient", "Message Summary", "Frequency", "Priority"],
    # Chemo related
    ["1", "Chemo session missed", "Email + SMS + Push", "Patient", "Emergency Contact + Doctor", "You missed your chemo session on [date]. Please contact your care team to reschedule immediately.", "Immediate + 24hr follow-up", "CRITICAL"],
    ["2", "Chemo session reminder (upcoming)", "Email + SMS + Push", "Patient", "Nurse", "Your next chemotherapy session is scheduled for [date] at [time]. Please arrive 30 min early for blood work.", "48hr + 24hr + 2hr before", "HIGH"],
    ["3", "Chemo cycle completed", "Email + Push", "Patient", "Doctor + Nurse", "Congratulations! You have completed cycle [X] of [Y]. Your next session is on [date].", "After each cycle", "MEDIUM"],
    ["4", "All chemo cycles completed", "Email + SMS", "Patient", "Doctor + Emergency Contact", "Your chemotherapy treatment is complete. Follow-up appointment scheduled for [date].", "Once", "HIGH"],
    # Doctor related
    ["5", "Doctor unavailable / on leave", "Email + SMS", "Patient", "Hospital Admin + Backup Doctor", "Dr. [Name] is unavailable on [date]. Your appointment has been reassigned to Dr. [Backup].", "Immediate", "HIGH"],
    ["6", "Doctor assignment change", "Email", "Patient", "New Doctor + Old Doctor", "Your oncologist has been changed to Dr. [Name]. Please contact the clinic for any questions.", "Immediate", "MEDIUM"],
    # Benchmark related
    ["7", "Tumor marker rising (above threshold)", "Email + Push", "Doctor", "Nurse", "ALERT: Patient [Name]'s [Marker] has increased from [X] to [Y]. Review and assess treatment plan.", "Immediate", "CRITICAL"],
    ["8", "Tumor marker improving", "Email + Push", "Patient", "Doctor", "Great news! Your [Marker] levels are trending downward. Keep following your treatment plan.", "After each result", "LOW"],
    ["9", "Benchmark deviation", "Email", "Doctor", "Hospital Admin", "Patient [Name]'s treatment response is below expected benchmarks for [Cancer Type] Stage [X].", "Weekly digest", "HIGH"],
    # ER related
    ["10", "ER visit logged", "Email + SMS", "Oncologist", "Nurse + Hospital Admin", "ALERT: Patient [Name] was admitted to ER on [date]. Reason: [Category]. Review medications.", "Immediate", "CRITICAL"],
    ["11", "Extra medication administered", "Email + Push", "Oncologist", "Chemist/Pharmacist", "Additional medication [Drug] was administered to Patient [Name] during ER visit. Review for interactions.", "Immediate", "HIGH"],
    ["12", "Unplanned hospitalization", "Email + SMS", "Doctor + Nurse", "Emergency Contact", "Patient [Name] has been hospitalized unexpectedly. Reason: [Details]. Oncology team notified.", "Immediate", "CRITICAL"],
    # AI-specific notifications
    ["13", "AI analysis complete", "Email + Push", "Doctor", "Nurse", "AI analysis of [Patient Name]'s [Report Type] is complete. Review extracted data and approve/reject.", "Immediate", "HIGH"],
    ["14", "AI detected critical finding", "Email + SMS + Push", "Oncologist", "Hospital Admin", "URGENT: AI detected [Stage IV / critical abnormality] in [Patient Name]'s pathology report. Immediate review required.", "Immediate", "CRITICAL"],
    ["15", "AI oncology record created (pending confirmation)", "Email + Push", "Doctor", "Nurse", "AI has auto-created an oncology record for [Patient Name] from pathology report. Doctor confirmation required.", "Immediate", "HIGH"],
    ["16", "AI treatment roadmap generated", "Email + Push", "Doctor", "Patient", "AI has generated a treatment roadmap for [Patient Name]. Review recommendations and customize treatment plan.", "Immediate", "MEDIUM"],
    ["17", "AI processing failed", "Email", "Hospital Admin", "Doctor", "AI processing failed for [Patient Name]'s report: [error]. Please retry or manually enter data.", "Immediate", "MEDIUM"],
    # Follow-up
    ["18", "Lab work / blood test reminder", "Email + SMS", "Patient", "Nurse", "Your blood work is due before your next chemo session on [date]. Please visit the lab by [deadline].", "72hr + 24hr before", "MEDIUM"],
    ["19", "Follow-up scan due", "Email + SMS + Push", "Patient", "Doctor", "Your follow-up [scan type] is due. Please schedule your appointment before [date].", "1 week + 3 days before", "MEDIUM"],
    ["20", "Treatment plan review due", "Email", "Doctor", "Nurse", "Patient [Name]'s treatment plan is due for review (last reviewed: [date]).", "Monthly or per protocol", "MEDIUM"],
    # System
    ["21", "Patient registered / onboarded", "Email", "Patient", "Doctor + Nurse", "Welcome to [Hospital] Oncology Care. Your treatment team and portal access details are ready.", "Once", "LOW"],
    ["22", "Insurance expiration warning", "Email + SMS", "Patient", "Hospital Admin", "Your insurance coverage is expiring on [date]. Please update your insurance information.", "30 days + 7 days before", "MEDIUM"],
    ["23", "Emergency contact not set", "Push + Email", "Patient", "Nurse", "Please add an emergency contact to your profile. This is required for your safety during treatment.", "Weekly until resolved", "HIGH"],
]

for r, row in enumerate(notif_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws7.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = ORANGE_FILL
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL
            if c == 8:
                if val == "CRITICAL":
                    cell.fill = RED_FILL
                    cell.font = Font(name="Calibri", bold=True, color="991B1B")
                elif val == "HIGH":
                    cell.fill = YELLOW_FILL
                    cell.font = Font(name="Calibri", bold=True, color="92400E")
                elif val == "MEDIUM":
                    cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
                elif val == "LOW":
                    cell.fill = GREEN_FILL

ws7.column_dimensions["A"].width = 6
ws7.column_dimensions["B"].width = 42
ws7.column_dimensions["C"].width = 22
ws7.column_dimensions["D"].width = 22
ws7.column_dimensions["E"].width = 30
ws7.column_dimensions["F"].width = 65
ws7.column_dimensions["G"].width = 28
ws7.column_dimensions["H"].width = 14


# ═══════════════════════════════════════════════════════════════════
# TAB 8 — BENCHMARKS & ANALYTICS
# ═══════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("8. Benchmarks & Analytics")
ws8.sheet_properties.tabColor = "0D9488"

bench_data = [
    ["Dashboard Section", "Metric / KPI", "Visualization", "Data Source", "Update Frequency", "Who Sees It", "Benchmark Comparison"],
    ["PATIENT DASHBOARD", "", "", "", "", "", ""],
    ["Treatment Progress", "Chemo cycles completed (X of Y)", "Progress bar + counter", "Chemo schedule module", "After each session", "Patient, Doctor, Nurse", "Expected vs Actual timeline"],
    ["Treatment Progress", "Days since last treatment", "Counter + status badge", "Chemo schedule module", "Real-time", "Patient, Doctor", "Protocol-defined interval"],
    ["Treatment Progress", "Next session countdown", "Calendar card + timer", "Chemo schedule module", "Real-time", "Patient, Nurse", "—"],
    ["Tumor Markers", "Marker value over time", "Line chart with trend arrow (↑↓→)", "Lab results / AI-extracted", "After each lab result", "Patient, Doctor", "Normal range reference lines"],
    ["Tumor Markers", "Marker change % (since last test)", "KPI card with color coding", "Lab results / AI-extracted", "After each lab result", "Patient, Doctor, Nurse", "Expected response range"],
    ["AI Insights", "AI confidence score for extraction", "Badge (High/Medium/Low)", "AI Analysis Service", "Per document", "Doctor, Admin", "—"],
    ["AI Insights", "AI-recommended treatment vs actual", "Comparison card", "AI roadmap + treatment records", "Per record", "Doctor", "NCCN guideline adherence"],
    ["Side Effects", "Side effect severity score", "Radar/spider chart", "Patient self-reporting", "Weekly", "Patient, Doctor, Nurse", "Average for same protocol"],
    ["Medications", "Medication adherence %", "Gauge chart", "Prescription module", "Daily", "Patient, Nurse, Chemist", "Target: >95%"],
    ["", "", "", "", "", "", ""],
    ["HOSPITAL DASHBOARD", "", "", "", "", "", ""],
    ["Chemo Operations", "Total active patients per cancer type", "Donut chart by cancer type", "Patient registry", "Real-time", "SuperAdmin, Hospital Admin", "Capacity benchmarks"],
    ["Chemo Operations", "Chemo completion rate %", "Bar chart per cancer type", "Chemo schedule module", "Weekly", "Admin, Doctor", "National average"],
    ["Chemo Operations", "Missed session rate %", "Trend line chart", "Chemo schedule module", "Weekly", "Admin, Doctor, Nurse", "Hospital target (<5%)"],
    ["Chemo Operations", "Average delay between cycles (days)", "Box plot", "Chemo schedule module", "Monthly", "Admin, Doctor", "Protocol-defined intervals"],
    ["Treatment Outcomes", "Treatment response rates (CR/PR/SD/PD)", "Stacked bar per cancer type", "Oncology assessments", "Monthly", "Admin, Doctor", "Published clinical trial rates"],
    ["Treatment Outcomes", "Recurrence rate %", "Trend line per cancer type", "Follow-up records", "Quarterly", "Admin, Doctor", "National cancer database"],
    ["AI Performance", "AI extraction accuracy %", "KPI card + trend", "AI vs manual comparison", "Monthly", "Admin", "Target: >95%"],
    ["AI Performance", "Average document processing time", "Bar chart by doc type", "Processing progress records", "Weekly", "Admin", "Target: <30 seconds"],
    ["AI Performance", "AI auto-created records confirmed %", "Gauge chart", "Oncology records audit", "Monthly", "Admin, Doctor", "Target: >90%"],
    ["ER & Emergency", "ER visits per 100 chemo cycles", "Bar chart + trend", "ER visit logs", "Monthly", "Admin, Doctor", "National oncology ER benchmark"],
    ["ER & Emergency", "Top ER visit reasons", "Pie chart / treemap", "ER visit logs", "Monthly", "Admin, Doctor, Nurse", "—"],
    ["ER & Emergency", "Unplanned hospitalization rate", "KPI card with trend", "ER visit logs", "Monthly", "Admin", "CMS benchmark"],
    ["Financial Metrics", "Cost per chemo cycle (avg)", "Bar chart per cancer type", "Billing module", "Monthly", "Admin", "Regional average"],
    ["Staff Performance", "Patients per oncologist", "Bar chart", "Assignment records", "Real-time", "Admin", "Recommended ratio"],
]

for r, row in enumerate(bench_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws8.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = TEAL_FILL
        elif val in ("PATIENT DASHBOARD", "HOSPITAL DASHBOARD"):
            cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            for cc in range(2, 8):
                ws8.cell(row=r, column=cc).fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL

ws8.column_dimensions["A"].width = 24
ws8.column_dimensions["B"].width = 42
ws8.column_dimensions["C"].width = 30
ws8.column_dimensions["D"].width = 28
ws8.column_dimensions["E"].width = 20
ws8.column_dimensions["F"].width = 30
ws8.column_dimensions["G"].width = 32


# ═══════════════════════════════════════════════════════════════════
# TAB 9 — ER & EMERGENCY TRACKING
# ═══════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("9. ER & Emergency Tracking")
ws9.sheet_properties.tabColor = "B91C1C"

er_data = [
    ["Category", "Field / Feature", "Description", "Captured By", "Visible To", "Triggers Notification?"],
    ["ER VISIT LOGGING", "", "", "", "", ""],
    ["Visit Details", "Visit Date & Time", "Date/time patient arrived at ER", "Nurse / ER Staff", "Doctor, Nurse, Admin", "Yes — Oncologist notified"],
    ["Visit Details", "ER Facility", "Which hospital/ER location", "Nurse / ER Staff", "Doctor, Admin", "No"],
    ["Visit Details", "Reason Category", "Dropdown: Fever, Infection, Pain, Chemo Reaction, Bleeding, Breathing Difficulty, Fall, Other", "Nurse / ER Staff", "Doctor, Nurse, Admin, Patient", "Yes — if chemo-related"],
    ["Visit Details", "Detailed Description", "Free-text clinical notes about the visit", "Doctor / Nurse", "Doctor, Nurse", "No"],
    ["Visit Details", "Severity Level", "Triage: Critical / Urgent / Standard", "ER Staff", "All Care Team", "Yes — if Critical"],
    ["Visit Details", "Outcome", "Treated & Released / Admitted / Transferred / Against Medical Advice", "Doctor", "All Care Team + Admin", "Yes — if Admitted"],
    ["Visit Details", "Duration of Stay", "Hours in ER / Days if hospitalized", "System (calculated)", "Doctor, Admin", "No"],
    ["", "", "", "", "", ""],
    ["EMERGENCY MEDICATIONS", "", "", "", "", ""],
    ["Extra Meds", "Medication Name", "Drug administered during ER visit", "Doctor / Nurse", "Doctor, Nurse, Chemist", "Yes — Pharmacist alert"],
    ["Extra Meds", "Dosage & Route", "Amount + IV/Oral/IM/SC", "Doctor / Nurse", "Doctor, Nurse, Chemist", "No"],
    ["Extra Meds", "Reason for Extra Med", "Why this med was needed (linked to ER reason)", "Doctor", "Doctor, Chemist", "No"],
    ["Extra Meds", "Drug Interaction Check", "Auto-check against current chemo + support meds", "System (auto)", "Doctor, Chemist", "Yes — if interaction found"],
    ["Extra Meds", "Impact on Treatment Plan", "Chemo delayed / Protocol modified / No change", "Doctor", "Doctor, Nurse, Admin", "Yes — if plan changed"],
    ["", "", "", "", "", ""],
    ["FOLLOW-UP ACTIONS", "", "", "", "", ""],
    ["Follow-up", "Post-ER Assessment Required", "Flag: Does oncologist need to reassess?", "ER Doctor", "Oncologist", "Yes"],
    ["Follow-up", "Chemo Schedule Impact", "Next session: On time / Delayed / Cancelled", "Oncologist", "Patient, Nurse", "Yes — Patient notified"],
    ["Follow-up", "Additional Tests Ordered", "Labs, imaging, or consults ordered post-ER", "Doctor", "Nurse, Patient", "Yes — scheduling"],
    ["Follow-up", "Emergency Contact Notified", "Was emergency contact called? When?", "Nurse", "Admin", "Auto-logged"],
    ["", "", "", "", "", ""],
    ["ANALYTICS (Dashboard)", "", "", "", "", ""],
    ["Analytics", "ER visits per patient (over time)", "Track frequency of ER visits per patient", "System", "Doctor, Admin", "—"],
    ["Analytics", "ER visits by cancer type", "Compare ER rates across Breast/Prostate/Lung/Colorectal", "System", "Admin", "—"],
    ["Analytics", "Most common ER reasons", "Top reasons by category", "System", "Admin, Doctor", "—"],
    ["Analytics", "ER visits post-chemo (within 7 days)", "How many ER visits occur shortly after chemo", "System", "Admin, Doctor", "—"],
    ["Analytics", "Extra medication frequency", "Most commonly administered emergency drugs", "System", "Admin, Chemist", "—"],
    ["Analytics", "Cost per ER episode", "Average cost of ER visits per cancer type", "Billing", "Admin", "—"],
]

for r, row in enumerate(er_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws9.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="B91C1C", end_color="B91C1C", fill_type="solid")
        elif val in ("ER VISIT LOGGING", "EMERGENCY MEDICATIONS", "FOLLOW-UP ACTIONS", "ANALYTICS (Dashboard)"):
            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
            for cc in range(2, 7):
                ws9.cell(row=r, column=cc).fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL

ws9.column_dimensions["A"].width = 24
ws9.column_dimensions["B"].width = 35
ws9.column_dimensions["C"].width = 55
ws9.column_dimensions["D"].width = 22
ws9.column_dimensions["E"].width = 30
ws9.column_dimensions["F"].width = 25


# ═══════════════════════════════════════════════════════════════════
# TAB 10 — TECH STACK & ARCHITECTURE
# ═══════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("10. Tech Stack")
ws10.sheet_properties.tabColor = "6366F1"

tech_data = [
    ["Layer", "Technology", "Purpose", "Notes"],
    ["Frontend", "React 18 + TypeScript", "Patient portal & Hospital dashboard SPA", "Vite build tool, responsive design"],
    ["Frontend", "Tailwind CSS / Ant Design", "UI component library & styling", "Consistent medical-grade UI"],
    ["Frontend", "React Query (TanStack)", "API state management & caching", "Optimistic updates, real-time sync"],
    ["Frontend", "Chart.js / Recharts", "Benchmark visualizations & trend charts", "Tumor marker graphs, analytics"],
    ["Frontend", "React Router v6", "Client-side routing & role-based routes", "Protected routes per role"],
    ["", "", "", ""],
    ["Backend", "NestJS + TypeScript", "REST API server & business logic", "Modular architecture, guards, interceptors"],
    ["Backend", "TypeORM", "Database ORM & migrations", "Entity-based models, relations"],
    ["Backend", "PostgreSQL", "Primary relational database", "HIPAA-compliant data storage"],
    ["Backend", "Redis", "Caching, session store, rate limiting, job queue", "BullMQ backing for async processing"],
    ["Backend", "JWT + Passport", "Authentication & authorization", "Role-based guards, refresh tokens"],
    ["Backend", "class-validator + class-transformer", "DTO validation layer", "Request validation, whitelist, transform"],
    ["Backend", "Winston / Pino", "Structured logging", "Request logging, error tracking, audit trail"],
    ["", "", "", ""],
    ["AI / ML", "Groq API (Llama 3.3 70B)", "LLM-powered medical text analysis", "groq-sdk npm package, JSON mode"],
    ["AI / ML", "Tesseract OCR", "Text extraction from scanned documents", "node-tesseract-ocr + tesseract-ocr system binary"],
    ["AI / ML", "pdf-parse / pdfjs-dist", "PDF text extraction (digital PDFs)", "Extract text layer from born-digital PDFs"],
    ["AI / ML", "pdf-poppler / pdf2pic", "PDF to image conversion", "For OCR pipeline on scanned PDFs"],
    ["AI / ML", "sharp", "Image preprocessing for OCR", "Grayscale, contrast, sharpen, resize"],
    ["AI / ML", "AJCC Staging Engine", "Deterministic cancer staging (no AI)", "Pure TypeScript lookup tables, AJCC 8th Edition"],
    ["AI / ML", "BullMQ", "Async document processing queue", "Background processing with retries & priority"],
    ["", "", "", ""],
    ["Notifications", "SendGrid / AWS SES", "Email notification delivery", "Templates, tracking, compliance"],
    ["Notifications", "Twilio", "SMS notification delivery", "Patient reminders, emergency alerts"],
    ["Notifications", "Firebase Cloud Messaging", "Push notifications", "Mobile + browser push"],
    ["Notifications", "BullMQ (Redis)", "Async notification processing", "Retry logic, scheduled delivery"],
    ["", "", "", ""],
    ["Infrastructure", "Docker + Docker Compose", "Containerization", "Dev/staging/production parity"],
    ["Infrastructure", "AWS / GCP", "Cloud hosting", "HIPAA-eligible services"],
    ["Infrastructure", "GitHub Actions", "CI/CD pipeline", "Automated testing + deployment"],
    ["Infrastructure", "Nginx", "Reverse proxy & load balancer", "SSL termination, rate limiting"],
    ["", "", "", ""],
    ["Security", "HIPAA Compliance", "Healthcare data protection", "Encryption at rest + transit, audit logs"],
    ["Security", "RBAC (Role-Based Access)", "Permission enforcement", "6 roles with granular permissions"],
    ["Security", "Data Encryption", "AES-256 at rest, TLS 1.3 in transit", "PHI protection"],
    ["Security", "Audit Logging", "Track all data access & changes", "Compliance reporting"],
    ["", "", "", ""],
    ["Testing", "Jest", "Unit & integration testing", "Backend + Frontend"],
    ["Testing", "Cypress / Playwright", "End-to-end testing", "Critical workflow coverage"],
    ["Testing", "k6 / Artillery", "Load & performance testing", "Chemo scheduling + AI processing stress tests"],
]

for r, row in enumerate(tech_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws10.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL

ws10.column_dimensions["A"].width = 20
ws10.column_dimensions["B"].width = 36
ws10.column_dimensions["C"].width = 45
ws10.column_dimensions["D"].width = 45


# ═══════════════════════════════════════════════════════════════════
# TAB 11 — MILESTONES & DELIVERABLES
# ═══════════════════════════════════════════════════════════════════
ws11 = wb.create_sheet("11. Milestones")
ws11.sheet_properties.tabColor = "059669"

milestone_data = [
    ["#", "Milestone", "Target Date", "Deliverables", "Acceptance Criteria", "Dependencies", "Status"],

    ["M1", "Infrastructure & DevOps Complete", "Week 2",
     "• .env files for all environments\n• Docker + docker-compose working\n• CI/CD pipeline running\n• Health check endpoint\n• Logging configured\n• TypeORM migrations setup",
     "• All team members can run project via docker-compose up\n• CI builds pass on main branch\n• Migrations run cleanly\n• Logs structured and queryable",
     "None — first milestone", "Not Started"],

    ["M2", "DTOs & Validation Complete", "Week 4",
     "• DTOs created for ALL endpoints (30+ DTOs)\n• class-validator decorators on all fields\n• Swagger @ApiProperty on all DTOs\n• Global exception filter\n• Standardized API response format\n• File upload infrastructure (Multer + storage)",
     "• Invalid requests return proper 400 errors with field-level messages\n• Swagger docs show request/response schemas\n• File uploads working (PDF, images)\n• API responses follow consistent format",
     "M1 complete", "Not Started"],

    ["M3", "Auth & Role System Complete", "Week 6",
     "• RBAC with 6 roles + guards\n• Register/Forgot Password/Change Password\n• Audit logging on all auth events\n• Session management + token blacklisting\n• Rate limiting on auth endpoints\n• User management CRUD",
     "• Each role can only access permitted routes\n• JWT refresh + rotation working\n• Admin can CRUD users\n• Failed logins are rate-limited\n• Audit trail queryable",
     "M2 complete", "Not Started"],

    ["M4", "Patient Module & First Cancer Module", "Week 10",
     "• Patient registration with validation\n• Breast cancer diagnosis & chemo tracking\n• First cancer module fully functional\n• Patient portal v1 (read-only)",
     "• Doctor can create diagnosis + assign chemo\n• Patient sees treatment progress\n• Tumor markers tracked with trends\n• All inputs validated via DTOs",
     "M3 complete", "Not Started"],

    ["M5", "All 4 Cancer Modules Complete", "Week 16",
     "• Prostate, Lung, Colorectal modules\n• All chemo protocols implemented\n• Tumor marker tracking for each\n• Cancer-specific benchmarks",
     "• All 4 cancer types fully functional\n• Protocol-specific workflows working\n• Marker trend charts displaying",
     "M4 complete", "Not Started"],

    ["M6", "AI: OCR & Text Extraction Working", "Week 18",
     "• MedicalReport entity & file upload pipeline\n• PDF text extraction (digital + scanned)\n• Image OCR with preprocessing\n• Text cleanup & artifact removal\n• Document type auto-detection\n• Processing progress tracking",
     "• Can upload PDF/image → text extracted\n• Scanned PDFs OCR'd with >90% accuracy\n• Document type correctly classified\n• Progress shown in real-time",
     "M2 (file upload infrastructure)", "Not Started"],

    ["M7", "AI: LLM Analysis & Staging Working", "Week 22",
     "• Groq API integration (direct + chunked mode)\n• All 6 document-type-specific prompts\n• Abnormal lab detection\n• AJCC staging engine (all 4 cancer types)\n• Stage-specific recommendations\n• AI confidence scoring",
     "• Medical reports analyzed in <30s\n• Extracted data matches report content\n• AJCC staging correct per test suite\n• Large documents chunked & merged correctly\n• Abnormal labs flagged",
     "M6 complete", "Not Started"],

    ["M8", "AI: Auto-Population & Treatment Roadmap", "Week 24",
     "• Auto-create OncologyRecord from pathology\n• Treatment roadmap generation (NCCN-based)\n• Report approval → patient record population\n• Data source tracking (manual vs automated)\n• create-from-report API endpoint\n• Bulk document processing",
     "• Pathology report → OncologyRecord auto-created\n• Treatment roadmap has surgery/chemo/radiation/immuno/targeted recs\n• Approval populates vitals/meds/allergies/conditions\n• All auto-created records tagged data_source='automated'\n• Doctor can confirm/edit AI-created records",
     "M7 complete + M5 complete", "Not Started"],

    ["M9", "AI Processing Pipeline (End-to-End)", "Week 26",
     "• Full pipeline: upload → OCR → analyze → stage → populate\n• BullMQ background processing\n• WebSocket/SSE progress updates\n• Error recovery & retry\n• Queue management\n• Document complexity estimation",
     "• Complete pipeline runs in background\n• Frontend shows real-time progress\n• Failed documents can be retried\n• Concurrent processing works\n• Processing time estimated upfront",
     "M8 complete", "Not Started"],

    ["M10", "Chemo Engine & Medications", "Week 30",
     "• Chemo scheduling calendar\n• Missed session detection\n• Medication management\n• Chemist verification flow\n• Drug interaction alerts",
     "• Sessions auto-schedule per protocol\n• Missed sessions flagged within 24hr\n• Drug interactions detected\n• Chemist can verify/dispense",
     "M5 complete", "Not Started"],

    ["M11", "Notification System Live", "Week 32",
     "• Email + SMS + Push working\n• All 23 notification triggers active\n• AI-specific notifications (analysis complete, critical finding, etc.)\n• Alert rules engine\n• User preference management",
     "• Missed chemo triggers notification\n• AI critical finding sends urgent alert\n• Doctor unavailable alerts patient\n• AI processing complete notifies doctor",
     "M10 + M9 complete", "Not Started"],

    ["M12", "Dashboards & Patient Portal", "Week 36",
     "• Patient + Hospital dashboards\n• Tumor marker trend charts\n• AI performance metrics\n• ER analytics\n• Patient portal (treatment journey, AI summary, chemo counter)",
     "• All charts rendering correctly\n• AI extraction accuracy tracked\n• Role-based visibility enforced\n• Patient sees plain-language AI treatment summary",
     "M9 + M10 complete", "Not Started"],

    ["M13", "Reports, Compliance & UAT", "Week 38",
     "• PDF report generation\n• AI analysis audit trail\n• Insurance/billing exports\n• Compliance audit reports",
     "• Reports export correctly as PDF\n• AI audit logs complete and searchable\n• HIPAA compliance checklist passed",
     "M12 complete", "Not Started"],

    ["M14", "Testing & Production Launch", "Week 42",
     "• Full QA testing (>80% coverage)\n• E2E testing including AI pipeline\n• Security audit & pen testing\n• AI processing load testing\n• UAT with hospital staff\n• Production deployment\n• Documentation & training",
     "• >80% test coverage\n• Zero critical bugs\n• AI processing handles 50+ concurrent docs\n• Hospital staff trained\n• System stable under load\n• Full documentation delivered",
     "M1–M13 complete", "Not Started"],
]

for r, row in enumerate(milestone_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws11.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL
            # Highlight AI milestones
            if c == 2 and val and "AI" in str(val):
                cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
                cell.font = Font(name="Calibri", bold=True, size=11, color="3730A3")

ws11.column_dimensions["A"].width = 8
ws11.column_dimensions["B"].width = 38
ws11.column_dimensions["C"].width = 14
ws11.column_dimensions["D"].width = 48
ws11.column_dimensions["E"].width = 48
ws11.column_dimensions["F"].width = 32
ws11.column_dimensions["G"].width = 14
for r in range(2, len(milestone_data) + 1):
    ws11.row_dimensions[r].height = 110


# ═══════════════════════════════════════════════════════════════════
# TAB 12 — PATIENT PORTAL SCREENS
# ═══════════════════════════════════════════════════════════════════
ws12 = wb.create_sheet("12. Patient Portal Screens")
ws12.sheet_properties.tabColor = "8B5CF6"

portal_data = [
    ["Screen / Page", "Key Components", "Data Displayed", "Actions Available", "AI-Powered Features"],
    ["Dashboard (Home)", "• Treatment progress card\n• Next appointment card\n• Tumor marker mini-chart\n• Recent notifications\n• Quick actions\n• AI summary card",
     "• Chemo cycles: X of Y completed\n• Days until next session\n• Latest marker value with trend\n• Upcoming lab work dates",
     "• View full schedule\n• Contact care team\n• Report symptoms\n• View notifications",
     "• AI-generated treatment summary (plain language)\n• AI confidence badge on auto-extracted data\n• AI-detected trends highlighted"],
    ["My Treatment Plan", "• Cancer type & stage summary\n• Chemo protocol details\n• Cycle-by-cycle timeline\n• Medication list\n• AI treatment roadmap",
     "• Full treatment protocol\n• Each cycle: date, drugs, status\n• Completed ✓ / Upcoming / Missed ✗\n• Expected completion date",
     "• Download treatment summary\n• View cycle details\n• See medication info",
     "• AI treatment roadmap panel showing\n  recommended surgery/chemo/radiation/immuno\n• AJCC stage explanation card\n• Treatment intent (curative/palliative)"],
    ["Chemo Schedule", "• Calendar view (month/week)\n• Session detail cards\n• Pre-chemo checklist status\n• History log",
     "• All scheduled sessions\n• Session times & locations\n• Required pre-visit labs\n• Completed session summaries",
     "• View session details\n• See pre-chemo requirements\n• Download schedule",
     "• AI-suggested optimal scheduling\n• AI-flagged potential conflicts"],
    ["My Health Metrics", "• Tumor marker trend chart\n• Side effect tracker\n• Weight/vitals log\n• Treatment response indicator",
     "• Marker values over time (line chart)\n• Trend arrows (↑↓→)\n• Normal range reference\n• Side effect severity scores",
     "• Log symptoms\n• Log weight/vitals\n• View historical data\n• Download reports",
     "• AI-extracted lab results auto-populated\n• AI abnormal lab flagging (↑↓ arrows)\n• AI comparison to expected response curves"],
    ["My Medications", "• Current medication list\n• Dosage & schedule\n• Refill status\n• Side effect info",
     "• Chemo drugs (per cycle)\n• Supportive medications\n• Hormone therapy (if applicable)\n• Last dispensed dates",
     "• View drug information\n• Request refill\n• Report side effects",
     "• AI-extracted medications from reports\n• AI drug interaction warnings\n• Medications tagged: manual vs AI-extracted"],
    ["My Documents & Reports", "• Lab reports\n• Imaging reports\n• Pathology reports\n• Prescriptions\n• AI analysis results",
     "• Document list with dates\n• Document type tags\n• AI processing status badge\n• AI extracted summary",
     "• View/download documents\n• Upload documents\n• View AI analysis insights",
     "• AI extraction status per document\n• AI insights panel: key findings\n• AI recommendations panel\n• View what AI extracted vs original report"],
    ["Messages & Communication", "• Inbox (care team messages)\n• Compose new message\n• Contact directory\n• Urgent contact info",
     "• Messages from doctor/nurse\n• Automated system messages\n• AI processing notifications\n• Read/unread status",
     "• Reply to messages\n• New message to care team\n• Emergency call button",
     "• AI analysis completion notifications\n• AI-detected urgent finding alerts"],
    ["My Profile & Settings", "• Personal information\n• Emergency contacts\n• Insurance details\n• Notification preferences\n• Password/security",
     "• Demographics\n• Emergency contact list\n• Insurance status & expiry\n• Notification channel preferences",
     "• Edit profile\n• Add/edit emergency contacts\n• Update insurance\n• Toggle notification channels\n• Change password",
     "• AI-populated fields highlighted\n  (from report extraction)\n• Data source indicator:\n  'Entered by Doctor' vs 'AI-Extracted'"],
]

for r, row in enumerate(portal_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws12.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL

ws12.column_dimensions["A"].width = 26
ws12.column_dimensions["B"].width = 38
ws12.column_dimensions["C"].width = 38
ws12.column_dimensions["D"].width = 32
ws12.column_dimensions["E"].width = 38
for r in range(2, len(portal_data) + 1):
    ws12.row_dimensions[r].height = 130


# ═══════════════════════════════════════════════════════════════════
# TAB 13 — CURRENT STATE vs TARGET (Gap Analysis)
# ═══════════════════════════════════════════════════════════════════
ws13 = wb.create_sheet("13. Gap Analysis")
ws13.sheet_properties.tabColor = "F59E0B"

gap_data = [
    ["Area", "Django (Source)", "NestJS (Current State)", "Gap", "Action Required", "Phase"],

    ["INFRASTRUCTURE", "", "", "", "", ""],
    ["Environment Config", ".env with all settings", "Hardcoded defaults, no .env file", "CRITICAL GAP", "Create .env files for all environments", "Phase 1"],
    ["Migrations", "Django migrations (30+ files)", "TypeORM synchronize:true (no migrations)", "CRITICAL GAP", "Disable sync, generate initial migration", "Phase 1"],
    ["Docker", "N/A (manual deployment)", "N/A", "NEW — needed for both", "Create Dockerfiles + docker-compose", "Phase 1"],
    ["CI/CD", "N/A", "N/A", "NEW — needed for both", "Create GitHub Actions workflows", "Phase 1"],
    ["Logging", "Django debug logging", "No logging configured", "GAP", "Setup Winston/Pino structured logging", "Phase 1"],
    ["File Uploads", "Django media handling", "No file upload support", "CRITICAL GAP", "Configure Multer + storage service", "Phase 1"],

    ["", "", "", "", "", ""],
    ["VALIDATION", "", "", "", "", ""],
    ["DTOs", "Django REST serializers (full validation)", "ZERO DTOs — all endpoints use 'any' type", "CRITICAL GAP", "Create 30+ DTOs with class-validator", "Phase 1"],
    ["Error Handling", "Django REST error formatting", "Default NestJS exception handler", "GAP", "Custom exception filter + response format", "Phase 1"],
    ["Swagger Docs", "N/A (DRF browsable API)", "Swagger exists but minimal (no schemas)", "GAP", "Add @ApiProperty to all DTOs", "Phase 1"],

    ["", "", "", "", "", ""],
    ["AUTH & RBAC", "", "", "", "", ""],
    ["Authentication", "JWT (SimpleJWT) — login + refresh", "JWT (Passport) — login + refresh", "PARTIAL — same feature set", "Enhance: register, forgot password, audit", "Phase 2"],
    ["RBAC", "Django admin permissions", "No role guards — JWT only (all-or-nothing auth)", "CRITICAL GAP", "Implement @Roles() decorator + RolesGuard", "Phase 2"],
    ["User Management", "Django admin CRUD", "Only initial admin auto-seed", "GAP", "Full user CRUD endpoints per role", "Phase 2"],

    ["", "", "", "", "", ""],
    ["AI / ML", "", "", "", "", ""],
    ["MedicalReport Model", "Full model (15+ fields, file upload, status pipeline)", "NOT PORTED", "CRITICAL GAP", "Create MedicalReport entity + CRUD + upload", "Phase 4"],
    ["ProcessingProgress Model", "Progress tracking (steps, %, status)", "NOT PORTED", "GAP", "Create ProcessingProgress entity", "Phase 4"],
    ["Text Extraction (PDF)", "pdfplumber (digital PDFs)", "NOT PORTED", "CRITICAL GAP", "Use pdf-parse npm package", "Phase 4"],
    ["OCR (Scanned PDFs)", "pdf2image + pytesseract + PIL preprocessing", "NOT PORTED", "CRITICAL GAP", "Use pdf-poppler + node-tesseract-ocr + sharp", "Phase 4"],
    ["AI Analysis (Groq LLM)", "Groq API (Llama 3.3 70B) — 6 prompt types", "NOT PORTED", "CRITICAL GAP", "Use groq-sdk npm, port all 6 prompts", "Phase 4"],
    ["Document Chunking", "6000 token chunks, overlap, merge", "NOT PORTED", "GAP", "Port chunking + merge logic to TypeScript", "Phase 4"],
    ["Report Type Detection", "Keyword-based NLP classification", "NOT PORTED", "GAP", "Port keyword dictionaries + matching", "Phase 4"],
    ["Abnormal Lab Detection", "Reference range parsing + flagging", "NOT PORTED", "GAP", "Port parsing logic to TypeScript", "Phase 4"],
    ["AJCC Staging Engine", "581 lines — 4 cancer types", "NOT PORTED", "CRITICAL GAP", "Port lookup tables + decision trees", "Phase 4"],
    ["Treatment Roadmap (AI)", "2nd Groq API call for NCCN recommendations", "NOT PORTED", "CRITICAL GAP", "Port roadmap prompt + OncologyRecord fields", "Phase 4"],
    ["Auto Population", "Approval → populate vitals/meds/allergies", "NOT PORTED", "GAP", "Port DataPopulationService logic", "Phase 4"],
    ["Auto OncologyRecord", "Pathology → auto-create oncology record", "NOT PORTED", "CRITICAL GAP", "Port auto-creation + cancer type mapping", "Phase 4"],
    ["create_from_report API", "POST endpoint for processing reports", "NOT PORTED", "GAP", "Create endpoint + pipeline trigger", "Phase 4"],
    ["Background Processing", "Synchronous (in-request)", "NOT PORTED", "IMPROVEMENT", "Use BullMQ for async processing", "Phase 5"],

    ["", "", "", "", "", ""],
    ["EXISTING FEATURES (Already Ported)", "", "", "", "", ""],
    ["Patient CRUD", "Full REST API", "Full REST API ✓", "DONE", "Add DTOs for validation", "Phase 1"],
    ["Oncology CRUD", "Full REST API", "Full REST API ✓", "DONE", "Add DTOs for validation", "Phase 1"],
    ["Treatments/FollowUps/Symptoms", "Partial (missing ViewSets in Django)", "Full REST API ✓ (NestJS ahead)", "NestJS AHEAD", "No action needed", "—"],
    ["Payer Submissions", "Partial (missing ViewSet in Django)", "Full REST API ✓ (NestJS ahead)", "NestJS AHEAD", "No action needed", "—"],
    ["ChronicConditions", "Partial (missing ViewSet in Django)", "Full REST API ✓ (NestJS ahead)", "NestJS AHEAD", "No action needed", "—"],
    ["Dashboard Stats", "Custom admin dashboard", "GET /api/dashboard/stats ✓", "DONE", "Enhance with AI metrics", "Phase 8"],
    ["13 Entities", "13 Django models", "13 TypeORM entities ✓ (faithful mapping)", "DONE", "Add MedicalReport + ProcessingProgress", "Phase 4"],
]

for r, row in enumerate(gap_data, 1):
    for c, val in enumerate(row, 1):
        cell = ws13.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if r == 1:
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
        elif val in ("INFRASTRUCTURE", "VALIDATION", "AUTH & RBAC", "AI / ML",
                     "EXISTING FEATURES (Already Ported)"):
            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
            for cc in range(2, 7):
                ws13.cell(row=r, column=cc).fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        else:
            cell.font = BODY_FONT
            if r % 2 == 0:
                cell.fill = LIGHT_FILL
            # Color-code gap column
            if c == 4:
                if val == "CRITICAL GAP":
                    cell.fill = RED_FILL
                    cell.font = Font(name="Calibri", bold=True, color="991B1B")
                elif val == "GAP":
                    cell.fill = YELLOW_FILL
                    cell.font = Font(name="Calibri", bold=True, color="92400E")
                elif val == "DONE":
                    cell.fill = GREEN_FILL
                    cell.font = Font(name="Calibri", bold=True, color="065F46")
                elif val == "NestJS AHEAD":
                    cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, color="065F46")
                elif val == "PARTIAL":
                    cell.fill = YELLOW_FILL
                elif val == "IMPROVEMENT":
                    cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, color="3730A3")
                elif "NEW" in str(val):
                    cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")

ws13.column_dimensions["A"].width = 28
ws13.column_dimensions["B"].width = 42
ws13.column_dimensions["C"].width = 42
ws13.column_dimensions["D"].width = 18
ws13.column_dimensions["E"].width = 42
ws13.column_dimensions["F"].width = 14


# ═══════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════
output_path = "/home/tec/tpai/ehr_system_react_nest/Oncology_EHR_Project_Workplan.xlsx"
wb.save(output_path)
print(f"\n✅ Workplan v2 generated successfully!")
print(f"📄 File: {output_path}")
print(f"\n📊 Tabs created:")
print(f"   1.  Project Overview (with AI capabilities)")
print(f"   2.  Phase-Wise Plan (10 phases, 21 sprints, 42 weeks — INFRA FIRST)")
print(f"   3.  AI Features Detail (25+ AI features with Django→NestJS mapping)")
print(f"   4.  AI Tech Stack (Django→NestJS technology mapping)")
print(f"   5.  Roles & Permissions (6 roles, 65+ permissions including AI)")
print(f"   6.  Cancer Modules Detail (4 types with AI extraction + AJCC staging)")
print(f"   7.  Notification System (23 triggers including 5 AI-specific)")
print(f"   8.  Benchmarks & Analytics (30+ KPIs including AI performance)")
print(f"   9.  ER & Emergency Tracking")
print(f"  10.  Tech Stack (Frontend + Backend + AI/ML + Notifications + Infra)")
print(f"  11.  Milestones & Deliverables (14 milestones, 4 AI-specific)")
print(f"  12.  Patient Portal Screens (8 screens with AI features)")
print(f"  13.  Gap Analysis (Django vs NestJS current state + action items)")
print(f"\n🔑 KEY CHANGES from v1:")
print(f"   • Infrastructure & DevOps is now Phase 1 (DTOs, Docker, CI/CD, Config)")
print(f"   • AI Integration added as Phase 4-5 (OCR, LLM, Staging, Roadmap, Pipeline)")
print(f"   • New Tab 3: AI Features Detail — all 25+ AI features mapped")
print(f"   • New Tab 4: AI Tech Stack — Django→NestJS technology mapping")
print(f"   • New Tab 13: Gap Analysis — Django vs NestJS current state")
print(f"   • Roles tab includes AI document analysis permissions")
print(f"   • Notifications include 5 AI-specific triggers")
print(f"   • Benchmarks include AI performance metrics")
print(f"   • Patient Portal shows AI-powered features per screen")
print(f"   • 14 milestones (up from 10) — 4 are AI-specific")
